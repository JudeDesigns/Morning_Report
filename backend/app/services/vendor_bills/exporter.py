"""Generate 8-sheet vendor bills workbook (PRD 10.6 + Workflow v3 §10)."""
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from app.config import settings

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
LIGHT_RED_FILL = PatternFill("solid", fgColor="FFB3B3")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
PINK_FILL = PatternFill("solid", fgColor="FFC7CE")

# All Issues type-cell fills per Workflow v3 §10.3
ISSUE_MISSING_FILL = PatternFill("solid", fgColor="DDEBF7")
ISSUE_EXTRA_FILL = PatternFill("solid", fgColor="FCE4D6")
ISSUE_QTY_FILL = PatternFill("solid", fgColor="FFF2CC")

# Office tasks raw → display label (v3 §1.4)
OFFICE_TASK_TYPE_LABELS = {
    "order_header": "Order header",
    "item_task": "Item task",
    "embedded_task": "Item task",  # legacy value
}


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_po_bank_sheet(ws, po_rows: list[dict]):
    # Workflow v3 §3/§4.10/§17 — PO Bank retains only unprocessed rows;
    # processed POs are either reflected on working sheets or swept into
    # PO Items Not Charged at export time.
    ws.title = "Original PO - PO Bank"
    headers = ["Terms", "Ref #", "Date", "Vendor", "Memo", "Total", "PO Cost",
               "Description", "Item Code", "Qty", "Class", "Case Avg Wt", "Unit Avg Wt", "Status"]
    ws.append(headers)
    _style_header(ws)
    for row in po_rows:
        if row.get("status") == "processed":
            continue
        ws.append([
            row.get("terms"), row.get("ref_number"), row.get("txn_date"), row.get("vendor"),
            row.get("memo"), row.get("total_amount"), row.get("po_cost"),
            row.get("description"), row.get("item_code"), row.get("quantity"),
            row.get("class_name"), row.get("case_avg_weight"), row.get("unit_avg_weight"),
            row.get("status"),
        ])


def build_office_tasks_sheet(ws, tasks: list[dict]):
    # v3 §1.4 column layout
    ws.title = "Office and Driver's tasks"
    headers = ["Vendor", "Ref #", "Type", "Item", "Need Review", "Task / Instructions"]
    ws.append(headers)
    _style_header(ws)
    for task in tasks:
        type_raw = task.get("task_type") or ""
        ws.append([
            task.get("vendor_name"),
            task.get("ref_number"),
            OFFICE_TASK_TYPE_LABELS.get(type_raw, type_raw),
            task.get("item"),
            task.get("need_review"),
            task.get("task_instructions"),
        ])


def build_bill_import_sheet(ws, rows: list[dict]):
    ws.title = "Bill Import"
    headers = ["Line", "UPC", "Item", "Description", "Price", "Total", "Qty",
               "Ref", "Date", "Vendor", "Type"]
    ws.append(headers)
    _style_header(ws)
    for row in rows:
        row_idx = ws.max_row + 1
        ws.append([
            row.get("line"), row.get("upc"), row.get("item_code"), row.get("description"),
            row.get("price"), row.get("total"), row.get("qty"), row.get("ref"),
            row.get("date"), row.get("vendor"), row.get("type", "Inventory Part"),
        ])
        hl = row.get("highlight_status")
        if hl in ("not_on_po", "missing_po_item"):
            for col in range(1, 12):
                ws.cell(row_idx, col).fill = LIGHT_RED_FILL


def build_summary_vendor_sheet(ws, blocks: list[dict], items_by_block: dict):
    ws.title = "Summary Vendor & Warehouse"
    row_idx = 1
    for block in blocks:
        if not block.get("has_reportable_issue"):
            continue
        ws.cell(row_idx, 1, f"Vendor: {block.get('vendor')} | Invoice: {block.get('invoice_number')} | {block.get('invoice_date')}")
        ws.cell(row_idx, 1).font = Font(bold=True)
        row_idx += 1
        ws.cell(row_idx, 1, "EMAIL SUBJECT:")
        ws.cell(row_idx, 2, block.get("email_subject", ""))
        row_idx += 1
        ws.cell(row_idx, 1, "--- COPY/PASTE EMAIL BODY ---")
        row_idx += 1
        for line in (block.get("email_body") or "").split("\n"):
            ws.cell(row_idx, 1, line)
            row_idx += 1

        items = items_by_block.get(str(block.get("id")), [])
        sections = {"overcharged": "OVERCHARGED ITEMS", "undercharged": "UNDERCHARGED / PRICE DECREASE ITEMS",
                    "qty_issue": "QUANTITY ISSUES", "missing": "MISSING / NOT BILLED ITEMS"}
        for sec_key, sec_title in sections.items():
            sec_items = [i for i in items if i.get("section") == sec_key]
            if not sec_items:
                continue
            ws.cell(row_idx, 1, sec_title)
            ws.cell(row_idx, 1).font = Font(bold=True)
            row_idx += 1
            ws.cell(row_idx, 1, "Item Code")
            ws.cell(row_idx, 2, "Description")
            ws.cell(row_idx, 3, "PO Rate")
            ws.cell(row_idx, 4, "Invoice Rate")
            ws.cell(row_idx, 5, "Qty")
            ws.cell(row_idx, 6, "Diff/Unit")
            ws.cell(row_idx, 7, "Total Impact")
            ws.cell(row_idx, 8, "Action")
            row_idx += 1
            for item in sec_items:
                ws.cell(row_idx, 1, item.get("item_code"))
                ws.cell(row_idx, 2, item.get("item_description"))
                ws.cell(row_idx, 3, item.get("po_rate"))
                ws.cell(row_idx, 4, item.get("invoice_rate"))
                ws.cell(row_idx, 5, item.get("qty"))
                ws.cell(row_idx, 6, item.get("difference_per_unit"))
                ws.cell(row_idx, 7, item.get("total_impact"))
                ws.cell(row_idx, 8, item.get("action_needed"))
                row_idx += 1
        row_idx += 1  # blank row between vendor blocks


def build_po_not_charged_sheet(ws, rows: list[dict]):
    ws.title = "PO Items Not Charged"
    if not rows:
        headers = ["Item Code", "Description", "Vendor", "Ref #", "Qty Ordered", "PO Cost", "ETA Request"]
        ws.append(headers)
        _style_header(ws)
        return
    headers = ["Item Code", "Description", "Vendor", "Ref #", "Qty Ordered", "PO Cost", "ETA Request"]
    ws.append(headers)
    _style_header(ws)
    for row in rows:
        ws.append([row.get("item_code"), row.get("description"), row.get("vendor"),
                   row.get("ref_number"), row.get("qty_ordered"), row.get("po_cost"),
                   "Yes" if row.get("eta_request") else ""])


def build_individual_weights_sheet(ws, weight_rows: list[dict]):
    ws.title = "Individual Weights"
    if not weight_rows:
        return
    # Build header dynamically based on max weights
    max_weights = max((len(r.get("weights") or [])) for r in weight_rows) if weight_rows else 0
    headers = ["Item Code (PO)", "Item Code (Bill)", "Product Name (PO)"]
    for i in range(1, max_weights + 1):
        headers.append(f"W{i}")
    headers.append("Total")
    ws.append(headers)
    _style_header(ws)
    for row in weight_rows:
        weights = row.get("weights") or []
        data = [row.get("item_code_po"), row.get("item_code_bill"), row.get("product_name_po")]
        data.extend(weights)
        data.extend([None] * (max_weights - len(weights)))
        data.append(row.get("total"))
        ws.append(data)


def build_cost_comparison_sheet(ws, rows: list[dict]):
    ws.title = "Cost Comparison"
    headers = ["Item Code", "Description", "Vendor", "PO Cost", "Vendor Bill Cost", "Difference", "% Change"]
    ws.append(headers)
    _style_header(ws)
    for row in rows:
        row_idx = ws.max_row + 1
        ws.append([row.get("item_code"), row.get("description"), row.get("vendor"),
                   row.get("po_cost"), row.get("vendor_bill_cost"),
                   row.get("difference"), row.get("percent_change")])
        diff = row.get("difference")
        if diff is not None:
            fill = GREEN_FILL if diff < 0 else (PINK_FILL if diff > 0 else None)
            if fill:
                for col in range(1, 8):
                    ws.cell(row_idx, col).fill = fill


def _fmt_money(v) -> str:
    try:
        return f"{float(v):.2f}"
    except (TypeError, ValueError):
        return "0.00"


def _fmt_qty(v) -> str:
    try:
        f = float(v)
        return f"{f:.0f}" if f.is_integer() else f"{f:.2f}"
    except (TypeError, ValueError):
        return str(v or "")


def _build_all_issues_rows(
    bill_import_rows: list[dict],
    summary_items_by_block: dict,
    summary_blocks: list[dict],
    po_not_charged: list[dict],
) -> list[dict]:
    """Assemble All Issues rows from existing data per Workflow v3 §10."""
    rows: list[dict] = []

    # ── Missing ──────────────────────────────────────────────────────────────
    for po in po_not_charged:
        qty = po.get("qty_ordered") or 0
        cost = po.get("po_cost") or 0
        amt = float(qty) * float(cost)
        inv_list = po.get("invoice_numbers") or []
        inv_display = ", ".join(inv_list) if inv_list else "(no related invoice)"
        ref = po.get("ref_number") or "—"
        detail = (
            f"Ordered on PO {ref} but not charged on invoice {inv_display}. "
            f"PO value ~${_fmt_money(amt)} ({_fmt_qty(qty)} @ ${_fmt_money(cost)})."
        )
        rows.append({
            "type": "Missing",
            "item_code": po.get("item_code") or "",
            "qty_ordered": _fmt_qty(qty),
            "description": po.get("description") or "",
            "vendor_ref": f"{po.get('vendor') or ''} · Inv {inv_display} · PO {ref}",
            "detail": detail,
            "_dollar_size": amt,
        })

    # ── Extra (billed not on PO) ─────────────────────────────────────────────
    for r in bill_import_rows:
        if r.get("highlight_status") != "not_on_po":
            continue
        qty = r.get("qty") or 0
        rate = r.get("price") or 0
        total = r.get("total") or (float(qty) * float(rate))
        inv = r.get("ref") or "—"
        detail = (
            f"Billed on invoice {inv} but not on PO. "
            f"${_fmt_money(total)} ({_fmt_qty(qty)} @ ${_fmt_money(rate)}). "
            "Confirm intended / provide PO."
        )
        rows.append({
            "type": "Extra",
            "item_code": "— (no PO code)",
            "qty_ordered": _fmt_qty(qty),
            # Description may include PO/vendor merge text on multiple lines;
            # take the last non-empty line as the vendor item label.
            "description": next(
                (ln.strip() for ln in reversed((r.get("description") or "").split("\n")) if ln.strip()),
                "",
            ),
            "vendor_ref": f"{r.get('vendor') or ''} · Inv {inv} · PO —",
            "detail": detail,
            "_dollar_size": float(total or 0),
        })

    # ── Qty mismatch ─────────────────────────────────────────────────────────
    # summary_items_by_block: { block_id: [item, …] }; each item already carries
    # vendor + invoice_number + po_ref_number + po_qty + bill_qty + qty_diff.
    for items in summary_items_by_block.values():
        for it in items:
            if not it.get("has_qty_mismatch"):
                continue
            billed = it.get("bill_qty") or 0
            ordered = it.get("po_qty") or 0
            delta = it.get("qty_diff") or (float(billed) - float(ordered))
            rate = it.get("invoice_rate") or 0
            inv = it.get("invoice_number") or "—"
            ref = it.get("po_ref_number") or "—"
            sign = "+" if delta >= 0 else ""
            detail = (
                f"Invoice {inv} billed {_fmt_qty(billed)}; "
                f"PO ordered {_fmt_qty(ordered)} ({sign}{_fmt_qty(delta)})."
            )
            # If the line also moved on rate, append a pointer (v3 §10.4).
            if it.get("section") in ("overcharged", "undercharged"):
                detail += " See Cost Comparison for rate move."
            rows.append({
                "type": "Qty mismatch",
                "item_code": it.get("item_code") or "",
                "qty_ordered": _fmt_qty(ordered),
                "description": it.get("item_description") or "",
                "vendor_ref": f"{it.get('vendor') or ''} · Inv {inv} · PO {ref}",
                "detail": detail,
                "_dollar_size": abs(float(delta)) * float(rate),
            })

    # Sort: group Missing → Extra → Qty mismatch; within each, dollar size desc
    type_order = {"Missing": 0, "Extra": 1, "Qty mismatch": 2}
    rows.sort(key=lambda r: (type_order.get(r["type"], 99), -r["_dollar_size"]))
    return rows


def build_all_issues_sheet(ws, rows: list[dict]):
    """Workflow v3 §10 — consolidated cross-invoice issues sheet."""
    ws.title = "All Issues"
    headers = ["Type", "Item Code", "Quantity Ordered", "Item Description",
               "Vendor (Invoice / PO#)", "Detail"]
    ws.append(headers)
    _style_header(ws)
    for r in rows:
        row_idx = ws.max_row + 1
        ws.append([
            r.get("type"), r.get("item_code"), r.get("qty_ordered"),
            r.get("description"), r.get("vendor_ref"), r.get("detail"),
        ])
        t = r.get("type")
        fill = (ISSUE_MISSING_FILL if t == "Missing"
                else ISSUE_EXTRA_FILL if t == "Extra"
                else ISSUE_QTY_FILL if t == "Qty mismatch"
                else None)
        if fill:
            ws.cell(row_idx, 1).fill = fill


def generate_workbook(run, po_rows, office_tasks, bill_import_rows, summary_blocks,
                      summary_items_by_block, po_not_charged, weight_rows, cost_comparison) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_po_bank_sheet(wb.create_sheet(), po_rows)
    build_office_tasks_sheet(wb.create_sheet(), office_tasks)
    build_bill_import_sheet(wb.create_sheet(), bill_import_rows)
    build_summary_vendor_sheet(wb.create_sheet(), summary_blocks, summary_items_by_block)
    # All Issues sits immediately after Summary Vendor & Warehouse (v3 §10)
    all_issues_rows = _build_all_issues_rows(
        bill_import_rows, summary_items_by_block, summary_blocks, po_not_charged,
    )
    build_all_issues_sheet(wb.create_sheet(), all_issues_rows)
    build_po_not_charged_sheet(wb.create_sheet(), po_not_charged)
    build_individual_weights_sheet(wb.create_sheet(), weight_rows)
    build_cost_comparison_sheet(wb.create_sheet(), cost_comparison)

    storage_root = Path(settings.STORAGE_PATH) / str(run.id) / "exports"
    storage_root.mkdir(parents=True, exist_ok=True)
    out_path = storage_root / f"QB_Vendor_Bill_Import_{run.id}.xlsx"
    wb.save(str(out_path))
    return str(out_path)
