"""Parse All Orders, Item List, Inventory, Shopping History sheets."""
from decimal import Decimal, InvalidOperation
from typing import Optional
import openpyxl
from openpyxl import load_workbook

# Unmapped All Orders columns that get passed through to the export verbatim.
# Mirrors PRD §8.3 which only names cols A, B, G, H, I, J, L, O, Q, S.
PASSTHROUGH_COL_INDICES = (2, 3, 4, 5, 10, 12, 13, 15, 17)  # C D E F K M N P R


def normalize_code(val) -> Optional[str]:
    """Normalize item code: strip whitespace, remove trailing .0, convert to str."""
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    # Remove trailing .0 from spreadsheet numeric strings
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def safe_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        return Decimal(str(val)).normalize()
    except (InvalidOperation, ValueError):
        return None


def _cell_passthrough(val):
    """Coerce a cell value into JSON-serialisable form for passthrough."""
    if val is None:
        return None
    if isinstance(val, (int, float, str, bool)):
        return val
    if isinstance(val, Decimal):
        return float(val)
    if hasattr(val, "isoformat"):
        return val.isoformat()
    return str(val)


def _find_sheet(wb: openpyxl.Workbook, predicates) -> Optional[object]:
    """Pick the first sheet matching any predicate, in priority order."""
    for predicate in predicates:
        for name in wb.sheetnames:
            if predicate(name.lower().strip()):
                return wb[name]
    return None


def parse_all_orders(wb: openpyxl.Workbook) -> tuple[list[dict], list[Optional[str]]]:
    """Parse All Orders sheet. Returns (rows, source_headers)."""
    ws = _find_sheet(wb, [
        lambda s: "all orders" in s,
        lambda s: "orders" in s,
    ]) or wb.active

    rows: list[dict] = []
    source_headers: list[Optional[str]] = []
    headers_found = False
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            rows.append({"_empty": True, "_raw": list(row)})
            continue
        if not headers_found:
            source_headers = [str(c).strip() if c is not None else None for c in row]
            headers_found = True
            continue  # skip header row
        raw = list(row)
        passthrough = {
            i: _cell_passthrough(raw[i]) if i < len(raw) else None
            for i in PASSTHROUGH_COL_INDICES
        }
        rows.append({
            "product_name": str(raw[0]).strip() if raw[0] else None,  # A
            "code": normalize_code(raw[1]),                            # B
            "qty": safe_decimal(raw[6]) if len(raw) > 6 else None,    # G
            "weight": safe_decimal(raw[7]) if len(raw) > 7 else None, # H
            "individual_weight_status": str(raw[8]).strip() if len(raw) > 8 and raw[8] else None,  # I
            "price": safe_decimal(raw[9]) if len(raw) > 9 else None,  # J
            "customer_name": str(raw[11]).strip() if len(raw) > 11 and raw[11] else None, # L
            "transaction_date": raw[14] if len(raw) > 14 else None,   # O
            "route": str(raw[16]).strip() if len(raw) > 16 and raw[16] else None, # Q
            "remark": str(raw[18]).strip() if len(raw) > 18 and raw[18] else None, # S
            "passthrough": passthrough,
            "_raw": raw,
        })
    return rows, source_headers


def parse_item_list(wb: openpyxl.Workbook) -> dict:
    """Parse Item List sheet. Returns dict keyed by normalized SKU CODE."""
    ws = _find_sheet(wb, [
        lambda s: "item" in s and "list" in s,
        lambda s: "item" in s and "master" in s,
        lambda s: s == "items" or s.startswith("item"),
        lambda s: "sku" in s,
    ])
    if ws is None:
        return {}

    result = {}
    header_row = None
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if header_row is None:
            if any(row):
                header_row = [str(c).strip().lower() if c else "" for c in row]
            continue
        raw = list(row)
        if not any(raw):
            continue

        def get_col(name_parts):
            for part in name_parts:
                for idx, h in enumerate(header_row):
                    if part in h:
                        return raw[idx] if idx < len(raw) else None
            return None

        sku = normalize_code(get_col(["sku", "code", "b"]) or (raw[1] if len(raw) > 1 else None))
        if not sku:
            continue
        result[sku] = {
            "category_name": get_col(["category"]),
            "current_cost_price": safe_decimal(get_col(["cost price", "cost"])),
            "current_selling_price": safe_decimal(get_col(["selling price", "sell"])),
            "unit": get_col(["unit"]),
        }
    return result


def parse_shopping_history(wb: openpyxl.Workbook) -> dict:
    """Parse Shopping History. Returns dict keyed by normalized item code.

    Tolerates two-row headers (e.g. "Unit" / "Price" split across rows).
    """
    ws = _find_sheet(wb, [
        lambda s: "shopping" in s and "history" in s,
        lambda s: "shopping" in s,
        lambda s: "history" in s,
    ])
    if ws is None:
        return {}

    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]

    # Pick the first two non-empty rows as header lines and merge them
    header_lines: list[list] = []
    data_start_idx = 0
    for idx, row in enumerate(all_rows):
        if not any(row):
            continue
        if len(header_lines) < 2 and _looks_like_header(row):
            header_lines.append(row)
            data_start_idx = idx + 1
            continue
        break

    if not header_lines:
        return {}

    width = max(len(r) for r in header_lines)
    composite = []
    for i in range(width):
        parts = []
        for line in header_lines:
            cell = line[i] if i < len(line) else None
            if cell is not None and str(cell).strip():
                parts.append(str(cell).strip())
        composite.append(" ".join(parts).lower())

    result = {}
    for row in all_rows[data_start_idx:]:
        if not any(row):
            continue
        raw = list(row)

        def get_col(name_parts):
            for part in name_parts:
                for idx, h in enumerate(composite):
                    if part in h:
                        return raw[idx] if idx < len(raw) else None
            return None

        # PRD §8.3: Item code is in source col B.
        item_code = normalize_code(get_col(["item"]) or (raw[1] if len(raw) > 1 else None))
        if not item_code:
            continue
        result[item_code] = {
            "shopping_product": get_col(["product"]),
            "new_bin": get_col(["bin"]),
            "unit_price": safe_decimal(get_col(["unit price"])),
            "case_price": safe_decimal(get_col(["case price"])),
        }
    return result


def _looks_like_header(row: list) -> bool:
    """A row looks like a header if its non-empty cells are mostly text labels."""
    non_empty = [c for c in row if c is not None and str(c).strip()]
    if not non_empty:
        return False
    text_count = sum(1 for c in non_empty if isinstance(c, str) and not str(c).replace(".", "").replace("-", "").isdigit())
    return text_count >= max(1, len(non_empty) // 2)


def parse_inventory(wb: openpyxl.Workbook) -> dict:
    """Parse Inventory sheet. Returns dict keyed by normalized item code."""
    ws = _find_sheet(wb, [
        lambda s: "inventory" in s,
        lambda s: "invent" in s,
        lambda s: "stock" in s,
    ])
    if ws is None:
        return {}

    result = {}
    header_row = None
    for row in ws.iter_rows(values_only=True):
        if header_row is None:
            if any(row):
                header_row = [str(c).strip().lower() if c else "" for c in row]
            continue
        raw = list(row)
        if not any(raw):
            continue

        def get_col(name_parts):
            for part in name_parts:
                for idx, h in enumerate(header_row):
                    if part in h:
                        return raw[idx] if idx < len(raw) else None
            return None

        item_code = normalize_code(get_col(["item"]) or (raw[0] if len(raw) > 0 else None))
        if not item_code:
            continue
        result[item_code] = {
            "quantity_on_hand": safe_decimal(get_col(["quantity on hand", "qty on hand", "on hand"])),
            "inventory_cost": safe_decimal(get_col(["cost"])),
            "case_avg_weight": safe_decimal(get_col(["case avg", "case_avg"])),
            "unit_avg_weight": safe_decimal(get_col(["unit avg", "unit_avg"])),
            "bin_internal": get_col(["bin internal", "bin"]),
        }
    return result
