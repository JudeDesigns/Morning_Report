"""End-to-end workflow tests — build synthetic workbooks in-memory, push them
through the full file-based pipeline (parse → process → JSON store → export),
and assert the intended PRD behaviour.

No real DB, no real network. Anthropic calls are mocked at the SDK boundary.
"""
import asyncio
import base64
import json
import os
import tempfile
from datetime import date
from decimal import Decimal
from io import BytesIO
from unittest.mock import MagicMock, patch

STORAGE = tempfile.mkdtemp()
os.environ["STORAGE_PATH"] = STORAGE
os.environ.setdefault("SECRET_KEY", "workflow-tests-secret-key-xyz")

import openpyxl
import pytest

from app.store import runs as run_store, files_meta as file_store, results as result_store
from app.services import web_orders as web_orders_svc
from app.services import jetro as jetro_svc
from app.services import vendor_bills as vendor_bills_svc
from app.services import combined_price as combined_svc


# ---------------------------------------------------------------------------
# Helpers — build in-memory XLSX files matching PRD column layouts
# ---------------------------------------------------------------------------
def _save_wb_to_run(wb: openpyxl.Workbook, run_id: str, file_type: str, name: str) -> dict:
    """Save a workbook to disk under the run folder and register file metadata."""
    folder = os.path.join(STORAGE, str(run_id), file_type)
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, name)
    wb.save(path)
    return file_store.add_file(
        run_id=run_id, file_type=file_type, original_filename=name,
        storage_path=path,
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _wo_all_orders_wb(rows: list[dict]) -> openpyxl.Workbook:
    """Web Orders 'All Orders' sheet — PRD layout (A=product, B=code, G=qty,
    H=weight, I=status, J=price, L=customer, O=date, Q=route, S=remark)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "All Orders"
    header = [None] * 20
    header[0] = "Product"; header[1] = "Code"; header[6] = "Qty"; header[7] = "Weight"
    header[8] = "Status"; header[9] = "Price"; header[11] = "Customer"
    header[14] = "Date"; header[16] = "Route"; header[18] = "Remark"
    ws.append(header)
    for r in rows:
        line = [None] * 20
        line[0] = r.get("product")
        line[1] = r.get("code")
        line[6] = r.get("qty")
        line[7] = r.get("weight")
        line[8] = r.get("status")
        line[9] = r.get("price")
        line[11] = r.get("customer")
        line[14] = r.get("date")
        line[16] = r.get("route")
        line[18] = r.get("remark")
        ws.append(line)
    return wb


def _wo_item_list_wb(items: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Item List"
    ws.append(["Category Name", "SKU Code", "Cost Price", "Selling Price", "Unit"])
    for it in items:
        ws.append([it.get("category"), it.get("code"), it.get("cost"),
                   it.get("price"), it.get("unit")])
    return wb


def _wo_inventory_wb(items: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(["Item Code", "Quantity on Hand", "Cost", "Case Avg Weight",
               "Unit Avg Weight", "Bin Internal"])
    for it in items:
        ws.append([it.get("code"), it.get("qoh"), it.get("cost"),
                   it.get("case_avg"), it.get("unit_avg"), it.get("bin")])
    return wb


def _wo_shopping_history_wb(items: list[dict]) -> openpyxl.Workbook:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Shopping History"
    ws.append(["Product", "Item Code", "Bin", "Unit Price", "Case Price"])
    for it in items:
        ws.append([it.get("product"), it.get("code"), it.get("bin"),
                   it.get("unit_price"), it.get("case_price")])
    return wb


def _jetro_source_wb(rows: list[dict]) -> openpyxl.Workbook:
    """Jetro source sheet — PRD cols: C=qty(2), D=product(3), E=code(4),
    O=customer(14), X=cost(23), Y=selling(24), AF=case_avg(31), AI=unit(34)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Jetro Source"
    header = [None] * 40
    header[2] = "Qty"; header[3] = "Product"; header[4] = "Code"
    header[14] = "Customer"; header[23] = "Cost"; header[24] = "Selling"
    header[31] = "Case Avg"; header[34] = "Unit"
    ws.append(header)
    for r in rows:
        line = [None] * 40
        line[2] = r.get("qty"); line[3] = r.get("product"); line[4] = r.get("code")
        line[14] = r.get("customer"); line[23] = r.get("cost")
        line[24] = r.get("selling"); line[31] = r.get("case_avg")
        line[34] = r.get("unit")
        ws.append(line)
    return wb


def _rd_invoice_wb(invoice_number: str, grand_total: float, lines: list[dict]) -> openpyxl.Workbook:
    """Restaurant Depot XLSX — preamble with Invoice # and Grand Total, then
    header at ~row 17 with cols A=line, B=upc, C=item, D=desc, E=price,
    F=cu, G=qty, H=total."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Invoice"
    # Pre-amble
    ws.append([f"Invoice # {invoice_number}"])  # row 1
    for _ in range(14):
        ws.append([""])
    ws.append(["Description column header check"])  # filler before header
    ws.append(["Line", "UPC", "Item", "Description", "Price", "C/U", "Qty", "Total"])
    for line in lines:
        ws.append([line.get("line"), line.get("upc"), line.get("item"),
                   line.get("desc"), line.get("price"), line.get("cu"),
                   line.get("qty"), line.get("total")])
    ws.append([None, None, None, "Grand Total", None, None, None, grand_total])
    return wb


def _qb_po_export_wb(rows: list[dict]) -> openpyxl.Workbook:
    """QuickBooks PO export — header row with 'Vendor' and 'TxnLine Description'.
    Cols: A=Terms, B=Ref, C=TxnDate, D=Vendor, E=Memo, F=Total, G=Cost,
    H=TxnLine Description, I=TxnLine Item, J=Qty, K=Class, L=CaseAvg, M=UnitAvg."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "PO Export"
    ws.append(["Terms", "Ref Number", "Txn Date", "Vendor", "Memo", "Total",
               "TxnLine Cost", "TxnLine Description", "TxnLine Item", "Qty",
               "Class", "Case Avg", "Unit Avg"])
    for r in rows:
        ws.append([r.get("terms"), r.get("ref"), r.get("txn_date"),
                   r.get("vendor"), r.get("memo"), r.get("total"),
                   r.get("cost"), r.get("desc"), r.get("item"),
                   r.get("qty"), r.get("class"), r.get("case_avg"),
                   r.get("unit_avg")])
    return wb


# ---------------------------------------------------------------------------
# Workflow A — Web Orders Check
# ---------------------------------------------------------------------------
def test_web_orders_full_pipeline_flags_problems_and_splits_dates():
    """PRD §8 — full pipeline: parse → enrich → split same-day/future →
    run problem checks → persist lines → export."""
    run_date = date(2025, 6, 15)
    run = run_store.create_run("web_orders_check", "WO E2E", run_date.isoformat())
    rid = run["id"]

    # 1 good row, 1 below-cost (problem), 1 LBS weight mismatch (problem),
    # 1 SUBSTITUTED (skipped), 1 MISSING (always problem), 1 future-dated
    all_orders = [
        {"product": "Good A", "code": "1001", "qty": 2, "weight": None,
         "status": "WITH", "price": 15.0, "customer": "Cust1",
         "date": run_date, "route": "R1", "remark": ""},
        {"product": "Below Cost", "code": "1002", "qty": 1, "weight": None,
         "status": "WITH", "price": 5.0, "customer": "Cust2",
         "date": run_date, "route": "R1", "remark": ""},
        {"product": "Beef AVG", "code": "1003", "qty": 2, "weight": 20.0,
         "status": "WITH", "price": 100.0, "customer": "Cust3",
         "date": run_date, "route": "R1", "remark": ""},
        {"product": "Sub Item", "code": "1004", "qty": 1, "weight": None,
         "status": "SUBSTITUTED", "price": 9.99, "customer": "Cust4",
         "date": run_date, "route": "R1", "remark": "sub for 1003"},
        {"product": "Missing Item", "code": "1005", "qty": 1, "weight": None,
         "status": "MISSING", "price": 10.0, "customer": "Cust5",
         "date": run_date, "route": "R1", "remark": ""},
        {"product": "Future Item", "code": "1001", "qty": 1, "weight": None,
         "status": "WITH", "price": 15.0, "customer": "Cust6",
         "date": date(2025, 6, 20), "route": "R2", "remark": ""},
    ]
    item_list = [
        {"category": "Dry", "code": "1001", "cost": 10.0, "price": 15.0, "unit": "EA"},
        {"category": "Dry", "code": "1002", "cost": 10.0, "price": 15.0, "unit": "EA"},
        {"category": "Meat", "code": "1003", "cost": 50.0, "price": 100.0, "unit": "LBS"},
        {"category": "Dry", "code": "1004", "cost": 5.0, "price": 9.99, "unit": "EA"},
        {"category": "Dry", "code": "1005", "cost": 5.0, "price": 10.0, "unit": "EA"},
    ]
    inventory = [
        {"code": "1001", "qoh": 100, "cost": 10.0, "case_avg": None, "unit_avg": None, "bin": "A1"},
        {"code": "1003", "qoh": 50, "cost": 50.0, "case_avg": 5.0, "unit_avg": None, "bin": "M1"},
    ]
    shopping = [
        {"product": "Good A Shopping", "code": "1001", "bin": "A1",
         "unit_price": 1.0, "case_price": 12.0},
    ]

    _save_wb_to_run(_wo_all_orders_wb(all_orders), rid, "web_orders_all_orders", "orders.xlsx")
    _save_wb_to_run(_wo_item_list_wb(item_list), rid, "web_orders_item_list", "items.xlsx")
    _save_wb_to_run(_wo_inventory_wb(inventory), rid, "web_orders_inventory", "inv.xlsx")
    _save_wb_to_run(_wo_shopping_history_wb(shopping), rid, "web_orders_shopping_history", "sh.xlsx")

    result = asyncio.run(web_orders_svc.process_run(rid))
    assert result["success"] is True, result
    assert result["future_rows"] == 1
    # 3 problematic: below-cost, LBS mismatch, MISSING. SUBSTITUTED is skipped.
    assert result["problematic_rows"] == 3

    lines = result_store.load(rid, "web_orders_lines", [])
    by_code = {l["code"]: l for l in lines if not l.get("is_future")}
    assert by_code["1001"]["is_problematic"] is False
    assert by_code["1002"]["is_problematic"] is True
    assert any("below cost" in r.lower() for r in by_code["1002"]["problem_reasons"])
    assert by_code["1003"]["is_problematic"] is True
    assert any("lbs weight mismatch" in r.lower() for r in by_code["1003"]["problem_reasons"])
    assert by_code["1004"]["is_problematic"] is False  # SUBSTITUTED → skipped
    assert by_code["1005"]["is_problematic"] is True
    assert "MISSING" in by_code["1005"]["problem_reasons"][0]

    assert run_store.get_run(rid)["status"] == "processed"
    out_path = asyncio.run(web_orders_svc.export_workbook(rid))
    assert os.path.exists(out_path)

    run_store.delete_run(rid)


# ---------------------------------------------------------------------------
# Workflow B — Jetro / Restaurant Depot Reconciliation
# ---------------------------------------------------------------------------
def test_jetro_full_pipeline_classifies_issues_and_price_changes():
    """PRD §9 — Jetro pipeline produces: integrity status per invoice,
    Missing/Extra/Qty mismatch issues, price update rows."""
    run = run_store.create_run("jetro_reconciliation", "Jetro E2E", "2025-06-01")
    rid = run["id"]

    jetro_rows = [
        {"qty": 5, "product": "ITEM A", "code": "10001", "customer": "Cust1",
         "cost": 10.0, "selling": 15.0, "case_avg": None, "unit": "EA"},
        {"qty": 2, "product": "ITEM B", "code": "10002", "customer": "Cust2",
         "cost": 20.0, "selling": 30.0, "case_avg": None, "unit": "EA"},
        # Missing — ordered but won't be billed
        {"qty": 3, "product": "ITEM C", "code": "10003", "customer": "Cust3",
         "cost": 5.0, "selling": 8.0, "case_avg": None, "unit": "EA"},
    ]
    # A matches qty=5 with cost rise; B has qty mismatch (3 vs 2);
    # D extra. Total = 55 + 60 + 7 = 122.
    inv_lines = [
        {"line": 1, "upc": "1111", "item": "10001", "desc": "ITEM A",
         "price": 11.0, "cu": "C", "qty": 5, "total": 55.0},
        {"line": 2, "upc": "2222", "item": "10002", "desc": "ITEM B",
         "price": 20.0, "cu": "C", "qty": 3, "total": 60.0},
        {"line": 3, "upc": "4444", "item": "10004", "desc": "ITEM D EXTRA",
         "price": 7.0, "cu": "C", "qty": 1, "total": 7.0},
    ]
    _save_wb_to_run(_jetro_source_wb(jetro_rows), rid, "jetro_source", "jetro.xlsx")
    _save_wb_to_run(_rd_invoice_wb("98765", 122.0, inv_lines), rid,
                    "restaurant_depot_invoice_xlsx", "inv.xlsx")

    result = asyncio.run(jetro_svc.process_run(rid))
    assert result["success"] is True, result
    assert result["invoices"] == 1
    invoices = result_store.load(rid, "jetro_invoices", [])
    assert invoices[0]["integrity_status"] == "ok"
    assert invoices[0]["invoice_number"] == "98765"

    issues = result_store.load(rid, "jetro_issues", [])
    issue_types = {i["issue_type"] for i in issues}
    assert {"Missing", "Extra", "Qty mismatch"} <= issue_types
    missing = [i for i in issues if i["issue_type"] == "Missing"]
    assert any(m["item_code"] == "10003" for m in missing)
    extras = [i for i in issues if i["issue_type"] == "Extra"]
    assert any(e["item_code"] == "10004" for e in extras)

    pus = result_store.load(rid, "jetro_price_updates", [])
    by_code = {p["item_code"]: p for p in pus}
    assert by_code["10001"]["old_cost"] == 10.0
    assert by_code["10001"]["new_cost"] == 11.0
    assert by_code["10001"]["cost_change_old_minus_new"] == -1.0

    out_path = asyncio.run(jetro_svc.export_workbook(rid))
    assert os.path.exists(out_path)

    run_store.delete_run(rid)


def test_jetro_integrity_mismatch_detected():
    """PRD §9.7 — if processed total ≠ printed grand total, mark mismatch."""
    run = run_store.create_run("jetro_reconciliation", "Jetro Mismatch", "2025-06-01")
    rid = run["id"]
    _save_wb_to_run(_jetro_source_wb([]), rid, "jetro_source", "j.xlsx")
    inv = [{"line": 1, "upc": "1", "item": "10001", "desc": "X",
            "price": 11.0, "cu": "C", "qty": 5, "total": 55.0}]
    _save_wb_to_run(_rd_invoice_wb("11111", 999.0, inv), rid,
                    "restaurant_depot_invoice_xlsx", "i.xlsx")

    asyncio.run(jetro_svc.process_run(rid))
    invoices = result_store.load(rid, "jetro_invoices", [])
    assert invoices[0]["integrity_status"] == "mismatch"
    run_store.delete_run(rid)


def _rd_invoice_wb_realistic(invoice_number: str, lines: list[dict],
                             sub_total: float, grand_total: float) -> openpyxl.Workbook:
    """Restaurant Depot XLSX that mirrors the exact shape of a real export:
    preamble row labeled 'Invoice' (col A) with the number in col C, header at
    row 17, Sub-Total and Grand Total both in column A, and trailer rows
    ('Total Cases', 'View A Quote') after Grand Total that must be ignored."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"
    for _ in range(6):
        ws.append([None])
    ws.append(["Invoice", None, int(invoice_number)])  # row 7
    for _ in range(9):
        ws.append([None])
    ws.append(["Line", "UPC", "Item", "Description", "Price", "C/U", "Qty", "Total", "T"])
    for line in lines:
        ws.append([line.get("line"), line.get("upc"), line.get("item"),
                   line.get("desc"), line.get("price"), line.get("cu"),
                   line.get("qty"), line.get("total"), None])
    # Banner/footer rows that previously polluted the output
    ws.append(["Delivery Charge", None, None, None, None, None, None, 0, None])
    ws.append(["Sub-Total", None, None, None, None, None, None, sub_total, None])
    ws.append([None])
    ws.append(["Grand Total", None, None, None, None, None, None, grand_total, None])
    ws.append(["Total Cases :  12", None, None, None, None, None, None, None, None])
    ws.append(["View All Quotes", None, None, "View A Quote", None, None, None, None, None])
    return wb


def test_jetro_handles_real_rd_invoice_shape():
    """PRD §9.4–9.9 — regression: real Restaurant Depot exports put
    'Sub-Total'/'Grand Total' in column A (not D), include trailer rows after
    Grand Total, and may contain banner rows like 'Delivery Charge' mid-data.
    Also covers: Missing aggregation per item code (PRD §9.9), LBS Qty mismatch
    reports ordered lbs (PRD §9.9 col C), and a coupon residue (qty=0 but
    total≠0) staying in the import so penny-integrity holds (PRD §9.8)."""
    run = run_store.create_run("jetro_reconciliation", "Jetro Realistic", "2025-06-01")
    rid = run["id"]

    # Jetro source: TWO customers ordering same missing item (must aggregate),
    # one LBS item with case_avg_weight, one regular item that will fold to
    # a coupon residue, one normal item.
    jetro_rows = [
        # Missing item, two customers
        {"qty": 3, "product": "MISSING ITEM", "code": "20001", "customer": "CustA",
         "cost": 4.0, "selling": 6.0, "case_avg": None, "unit": "EA"},
        {"qty": 2, "product": "MISSING ITEM", "code": "20001", "customer": "CustB",
         "cost": 4.0, "selling": 6.0, "case_avg": None, "unit": "EA"},
        # LBS item: ordered 3 cases × 15 lbs = 45 lbs, invoice will bill 50 lbs
        {"qty": 3, "product": "LBS ITEM", "code": "20002", "customer": "CustC",
         "cost": 5.0, "selling": 8.0, "case_avg": 15.0, "unit": "LBS"},
        # Normal billed item
        {"qty": 1, "product": "NORMAL", "code": "20003", "customer": "CustD",
         "cost": 10.0, "selling": 15.0, "case_avg": None, "unit": "EA"},
        # Item that ends up as qty=0/total=-1 after coupon + void in invoice
        {"qty": 1, "product": "COUPON RESIDUE", "code": "20004", "customer": "CustE",
         "cost": 20.0, "selling": 25.0, "case_avg": None, "unit": "EA"},
    ]

    # Invoice:
    #   20003 normal $10
    #   20002 LBS 50 lbs × $5 = $250
    #   20004 item $20 + coupon -$1 + void -$20  →  qty=0, total=-$1
    #   Extra item 20099 $7
    #   Sub-Total / Grand Total = 10 + 250 + (-1) + 7 = 266
    inv_lines = [
        {"line": 1, "upc": "1003", "item": "20003", "desc": "NORMAL",
         "price": 10.0, "cu": "C", "qty": 1, "total": 10.0},
        {"line": 2, "upc": "1002", "item": "20002", "desc": "LBS ITEM",
         "price": 5.0, "cu": "C", "qty": 50, "total": 250.0},
        {"line": 3, "upc": "1004", "item": "20004", "desc": "COUPON RESIDUE",
         "price": 20.0, "cu": "C", "qty": 1, "total": 20.0},
        {"line": 4, "upc": "Coupon", "item": "Coupon", "desc": "COUPON",
         "price": -1.0, "cu": "C", "qty": 1, "total": -1.0},
        {"line": 5, "upc": "1004", "item": "20004", "desc": "COUPON RESIDUE",
         "price": 20.0, "cu": "C", "qty": "1V", "total": -20.0},
        {"line": 6, "upc": "1099", "item": "20099", "desc": "EXTRA ITEM",
         "price": 7.0, "cu": "C", "qty": 1, "total": 7.0},
    ]

    _save_wb_to_run(_jetro_source_wb(jetro_rows), rid, "jetro_source", "j.xlsx")
    _save_wb_to_run(
        _rd_invoice_wb_realistic("230930", inv_lines, sub_total=266.0, grand_total=266.0),
        rid, "restaurant_depot_invoice_xlsx", "i.xlsx",
    )

    result = asyncio.run(jetro_svc.process_run(rid))
    assert result["success"] is True, result

    invoices = result_store.load(rid, "jetro_invoices", [])
    assert len(invoices) == 1
    inv = invoices[0]
    assert inv["invoice_number"] == "230930"
    assert inv["printed_grand_total"] == 266.0  # parser found Grand Total in col A
    assert inv["integrity_status"] == "ok"     # coupon residue preserved
    assert abs(inv["processed_total"] - 266.0) < 0.01

    # Trailer rows must NOT appear as items
    import_rows = result_store.load(rid, "jetro_import_rows", [])
    descs = [r["description"] for r in import_rows]
    assert "View A Quote" not in descs
    assert "Delivery Charge" not in descs
    assert "Sub-Total" not in descs
    assert "Grand Total" not in descs
    # All real items present, including the qty=0/total=-1 residue
    by_code = {r["item_code"]: r for r in import_rows}
    assert "20003" in by_code
    assert "20002" in by_code
    assert "20004" in by_code
    assert by_code["20004"]["qty"] == 0
    assert abs(by_code["20004"]["total"] - (-1.0)) < 0.001

    issues = result_store.load(rid, "jetro_issues", [])
    # Missing aggregation: 20001 ordered by two customers → exactly ONE row
    missing = [i for i in issues if i["issue_type"] == "Missing"]
    assert len(missing) == 1
    m = missing[0]
    assert m["item_code"] == "20001"
    assert m["quantity"] == 5  # 3 + 2 aggregated
    assert "CustA" in m["used_by"] and "CustB" in m["used_by"]
    assert m["dollar_size"] == 20.0  # cost 4 × qty 5

    # LBS Qty mismatch: column C must show ordered lbs (45), not billed (50)
    qmm = [i for i in issues if i["issue_type"] == "Qty mismatch" and i["item_code"] == "20002"]
    assert len(qmm) == 1
    assert qmm[0]["quantity"] == 45  # 3 cases × 15 lbs

    # Extra item 20099 must surface
    extras = [i for i in issues if i["issue_type"] == "Extra"]
    assert any(e["item_code"] == "20099" for e in extras)
    # No phantom Extra from trailer/banner rows
    assert not any(e["item_description"] in ("View A Quote", "Delivery Charge", "") for e in extras)

    run_store.delete_run(rid)


# ---------------------------------------------------------------------------
# Workflow C — Vendor Bills / PO Bank (with mocked Claude extraction)
# ---------------------------------------------------------------------------
def _fake_extraction_response(payload: dict) -> MagicMock:
    client = MagicMock()
    msg = MagicMock()
    msg.content = [MagicMock(text=json.dumps(payload))]
    client.messages.create.return_value = msg
    return client


PNG_BYTES = base64.b64decode(
    "iVBORw0KGgoAAAANSUhEUgAAAAEAAAABCAQAAAC1HAwCAAAAC0lEQVR4nGNgAAIAAAUAAeImBZsAAAAASUVORK5CYII="
)


def _write_bytes(rid: str, name: str, content: bytes) -> str:
    folder = os.path.join(STORAGE, str(rid), "vendor_bill_image")
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, name)
    with open(path, "wb") as f:
        f.write(content)
    return path


def test_vendor_bills_full_pipeline_with_image_extraction():
    """PRD §10 — full pipeline: parse QB PO export → extract bill image with
    Claude (mocked) → user-confirm lines → process bill → generate import rows,
    summary blocks, cost comparison. Then finalize sweeps unbilled POs."""
    run = run_store.create_run("vendor_bill_po_bank", "VB E2E", "2025-06-01")
    rid = run["id"]

    po_rows = [
        {"terms": "Net30", "ref": "PO-001", "txn_date": "2025-06-01",
         "vendor": "ACME Foods", "memo": "", "total": 200.0, "cost": 10.0,
         "desc": "TOMATO 25LB CASE", "item": "T-001", "qty": 10,
         "class": "Produce", "case_avg": None, "unit_avg": None},
        {"terms": "Net30", "ref": "PO-001", "txn_date": "2025-06-01",
         "vendor": "ACME Foods", "memo": "", "total": 50.0, "cost": 5.0,
         "desc": "ONION 50LB CASE", "item": "O-001", "qty": 10,
         "class": "Produce", "case_avg": None, "unit_avg": None},
        # Will not be billed → must appear in po_not_charged after finalize
        {"terms": "Net30", "ref": "PO-001", "txn_date": "2025-06-01",
         "vendor": "ACME Foods", "memo": "", "total": 30.0, "cost": 3.0,
         "desc": "GARLIC 5LB BAG", "item": "G-001", "qty": 10,
         "class": "Produce", "case_avg": None, "unit_avg": None},
    ]
    _save_wb_to_run(_qb_po_export_wb(po_rows), rid, "quickbooks_po_export", "po.xlsx")

    po_result = asyncio.run(vendor_bills_svc.process_po_upload(rid))
    assert po_result["success"] is True
    assert po_result["product_lines"] == 3

    # Upload a "bill image" and extract with mocked Claude.
    bill_file = file_store.add_file(
        run_id=rid, file_type="vendor_bill_image",
        original_filename="bill.png",
        storage_path=_write_bytes(rid, "bill.png", PNG_BYTES),
        mime_type="image/png",
    )
    extraction = {
        "vendor": "ACME Foods",
        "invoice_number": "INV-9001",
        "invoice_date": "2025-06-02",
        "bill_type": "invoice",
        "header_confidence": {"vendor": 0.99, "invoice_number": 0.99, "invoice_date": 0.99},
        "lines": [
            {"bill_item_code": "T-001", "description": "TOMATO 25LB CASE",
             "qty": 10, "rate": 10.0, "total": 100.0,
             "confidence": 0.97,
             "field_confidence": {"qty": 0.99, "rate": 0.99, "total": 0.99}},
            # Onion: overcharge $1/unit + low rate confidence
            {"bill_item_code": "O-001", "description": "ONION 50LB CASE",
             "qty": 10, "rate": 6.0, "total": 60.0,
             "confidence": 0.96,
             "field_confidence": {"qty": 0.99, "rate": 0.5, "total": 0.99}},
        ],
    }
    with patch("app.services.ai_extraction.settings.ANTHROPIC_API_KEY", "test-key"), \
         patch("app.services.ai_extraction.anthropic.Anthropic",
               return_value=_fake_extraction_response(extraction)):
        extract_result = asyncio.run(
            vendor_bills_svc.extract_bill_image(rid, bill_file["id"])
        )
    bill_id = extract_result["bill_id"]

    bills = result_store.load(rid, "vendor_bills", [])
    bill = next(b for b in bills if b["id"] == bill_id)
    assert bill["vendor_extracted"] == "ACME Foods"
    onion_line = next(l for l in bill["lines"] if l["bill_item_code"] == "O-001")
    assert "rate" in onion_line["field_needs_review"]

    # Simulate user review: confirm vendor + every line
    bill["vendor_confirmed"] = "ACME Foods"
    bill["extraction_status"] = "confirmed"
    for line in bill["lines"]:
        line["user_confirmed"] = True
    result_store.save(rid, "vendor_bills", bills)

    proc = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, bill_id))
    assert proc["success"] is True, proc
    assert proc["import_rows"] == 2
    assert proc["discrepancies"] == 1  # onion overcharged

    import_rows = result_store.load(rid, "bill_import_rows", [])
    assert {r["item_code"] for r in import_rows} == {"T-001", "O-001"}
    cc = result_store.load(rid, "cost_comparison", [])
    onion_cc = next(c for c in cc if c["item_code"] == "O-001")
    assert onion_cc["po_cost"] == 5.0
    assert onion_cc["vendor_bill_cost"] == 6.0
    assert onion_cc["difference"] == 1.0

    summary_blocks = result_store.load(rid, "vendor_summary_blocks", [])
    block = summary_blocks[0]
    assert block["vendor"] == "ACME Foods"
    assert block["invoice_number"] == "INV-9001"
    assert any(it["section"] == "overcharged" for it in block["items"])

    # PRD §10.12 — po_not_charged not eagerly populated; only after finalize.
    assert result_store.load(rid, "po_not_charged", []) == []
    fin = vendor_bills_svc.finalize_unbilled_po(rid)
    assert fin["added"] == 1
    pnc = result_store.load(rid, "po_not_charged", [])
    assert pnc[0]["item_code"] == "G-001"

    out_path = asyncio.run(vendor_bills_svc.export_workbook(rid))
    assert os.path.exists(out_path)

    run_store.delete_run(rid)


def test_vendor_bills_reopen_restores_swept_pos_for_late_bills():
    """Office signal: more bills arrived AFTER export. reopen_run must restore
    finalize-swept POs so a late bill can match against them, and the next
    finalize must only flag PO rows that are still unprocessed."""
    run = run_store.create_run("vendor_bill_po_bank", "VB Reopen", "2025-06-01")
    rid = run["id"]

    # PO Bank: two from same vendor, one already matched, one unmatched.
    result_store.save(rid, "po_bank", [
        {"id": "po-matched", "vendor": "ACME Foods", "item_code": "T-001",
         "description": "TOMATO", "ref_number": "PO-001",
         "quantity": 10, "po_cost": 10.0,
         "status": "processed", "processed_bill_id": "bill-1"},
        {"id": "po-leftover", "vendor": "ACME Foods", "item_code": "G-001",
         "description": "GARLIC", "ref_number": "PO-001",
         "quantity": 10, "po_cost": 3.0,
         "status": "unprocessed"},
    ])
    result_store.save(rid, "vendor_bills", [
        {"id": "bill-1", "vendor_confirmed": "ACME Foods",
         "vendor_extracted": "ACME Foods", "extraction_status": "processed",
         "invoice_number": "INV-1", "lines": []},
    ])

    # First finalize: garlic PO gets swept.
    fin = vendor_bills_svc.finalize_unbilled_po(rid)
    assert fin["added"] == 1
    po_bank = result_store.load(rid, "po_bank", [])
    leftover = next(p for p in po_bank if p["id"] == "po-leftover")
    assert leftover["status"] == "processed"
    assert leftover.get("processed_bill_id") in (None, "")
    assert len(result_store.load(rid, "po_not_charged", [])) == 1
    run_store.update_run(rid, {"status": "exported"})

    # Office uploads a late bill — reopen the run.
    outcome = vendor_bills_svc.reopen_run(rid)
    assert outcome["reverted_pos"] == 1
    po_bank = result_store.load(rid, "po_bank", [])
    leftover = next(p for p in po_bank if p["id"] == "po-leftover")
    assert leftover["status"] == "unprocessed"
    assert "processed_bill_id" not in leftover
    # Already-matched PO must NOT be touched.
    matched = next(p for p in po_bank if p["id"] == "po-matched")
    assert matched["status"] == "processed"
    assert matched["processed_bill_id"] == "bill-1"
    # po_not_charged cleared; run status reset.
    assert result_store.load(rid, "po_not_charged", []) == []
    assert run_store.get_run(rid)["status"] == "processed"

    # A new finalize after reopen flags the still-unmatched PO again.
    fin2 = vendor_bills_svc.finalize_unbilled_po(rid)
    assert fin2["added"] == 1

    run_store.delete_run(rid)


def test_vendor_bills_reopen_is_a_noop_when_no_pos_were_swept():
    """If finalize never ran (or no leftovers), reopen still resets status."""
    run = run_store.create_run("vendor_bill_po_bank", "VB Reopen Noop", "2025-06-01")
    rid = run["id"]
    result_store.save(rid, "po_bank", [
        {"id": "po-1", "vendor": "ACME", "status": "processed",
         "processed_bill_id": "bill-x"},
    ])
    run_store.update_run(rid, {"status": "exported"})
    outcome = vendor_bills_svc.reopen_run(rid)
    assert outcome["reverted_pos"] == 0
    assert run_store.get_run(rid)["status"] == "processed"
    run_store.delete_run(rid)


def test_vendor_bills_blocks_processing_with_unconfirmed_flagged_lines():
    """PRD §13 — lines with needs_review=True must be explicitly confirmed before processing.
    Lines with needs_review=False (high-confidence) are auto-approved and don't block."""
    run = run_store.create_run("vendor_bill_po_bank", "VB Unconfirmed", "2025-06-01")
    rid = run["id"]

    # Line with needs_review=True and user_confirmed=False → must block
    result_store.save(rid, "vendor_bills", [{
        "id": "b1", "vendor_confirmed": "X", "vendor_extracted": "X",
        "extraction_status": "review",
        "lines": [
            {"id": "l1", "bill_item_code": "A", "description": "A",
             "qty": 1, "rate": 1.0, "total": 1.0,
             "needs_review": True, "user_confirmed": False},
        ],
    }])
    result_store.save(rid, "po_bank", [])
    res = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, "b1"))
    assert res["success"] is False
    assert "flagged" in res["errors"][0]

    # Same line but needs_review=False → high-confidence, should pass through
    result_store.save(rid, "vendor_bills", [{
        "id": "b1", "vendor_confirmed": "X", "vendor_extracted": "X",
        "extraction_status": "review",
        "lines": [
            {"id": "l1", "bill_item_code": "A", "description": "A",
             "qty": 1, "rate": 1.0, "total": 1.0,
             "needs_review": False, "user_confirmed": False},
        ],
    }])
    result_store.save(rid, "po_bank", [])
    res2 = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, "b1"))
    assert res2["success"] is True, f"High-confidence line should not block: {res2}"

    run_store.delete_run(rid)


# ---------------------------------------------------------------------------
# Workflow D — Combined Price Changes
# ---------------------------------------------------------------------------
def test_combined_price_workflow_merges_jetro_and_vendor_sources():
    """PRD §11 — combined price changes pulls from a processed Jetro run AND a
    processed vendor-bill run, filters by NO_CHANGE threshold, sorts by |Δ| desc."""
    j_run = run_store.create_run("jetro_reconciliation", "J-src", "2025-06-01")
    jid = j_run["id"]
    result_store.save(jid, "jetro_price_updates", [
        {"item_code": "A1", "qty_charged": 10, "item_description": "Apple",
         "old_cost": 10.0, "new_cost": 12.0,
         "cost_change_old_minus_new": -2.0, "used_by": "Cust1"},
        # Below NO_CHANGE_THRESHOLD → must be filtered out
        {"item_code": "A2", "qty_charged": 5, "item_description": "Avo",
         "old_cost": 5.0, "new_cost": 5.001,
         "cost_change_old_minus_new": -0.001, "used_by": None},
    ])

    v_run = run_store.create_run("vendor_bill_po_bank", "V-src", "2025-06-01")
    vid = v_run["id"]
    result_store.save(vid, "cost_comparison", [
        {"item_code": "B1", "description": "Beef",
         "vendor": "ACME", "po_cost": 20.0, "vendor_bill_cost": 25.0,
         "difference": 5.0, "percent_change": 25.0},
    ])

    c_run = run_store.create_run("combined_price_changes", "Combined", "2025-06-01")
    cid = c_run["id"]
    result = asyncio.run(combined_svc.process_run(cid, jetro_run_id=jid, vendor_run_id=vid))
    assert result["success"] is True
    assert result["total_changes"] == 2  # A1 kept, A2 filtered, B1 kept
    rows = result_store.load(cid, "combined_price_rows", [])
    by_code = {r["item_code"]: r for r in rows}
    # Jetro 'difference' = new - old = +2 for A1
    assert by_code["A1"]["difference"] == 2.0
    assert by_code["A1"]["source"].startswith("Jetro")
    assert by_code["B1"]["difference"] == 5.0
    assert "ACME" in by_code["B1"]["source"]
    # Sorted by |Δ| desc
    assert rows[0]["item_code"] == "B1"
    assert rows[1]["item_code"] == "A1"

    out_path = asyncio.run(combined_svc.export_workbook(cid))
    assert os.path.exists(out_path)

    run_store.delete_run(jid)
    run_store.delete_run(vid)
    run_store.delete_run(cid)


def test_combined_price_with_no_source_runs_is_empty():
    run = run_store.create_run("combined_price_changes", "Empty", "2025-06-01")
    rid = run["id"]
    result = asyncio.run(combined_svc.process_run(rid, jetro_run_id=None, vendor_run_id=None))
    assert result["success"] is True
    assert result["total_changes"] == 0
    run_store.delete_run(rid)


def test_vendor_bills_catch_weight_qty_is_total_lbs_not_cases():
    """PRD §5 Bill Import — for weight-based items, the Bill Import Qty must be
    the total lbs shipped (so price × qty = total in QuickBooks), not the case
    count. Catch-weight lines are identified by AI-populated individual_weights."""
    run = run_store.create_run("vendor_bill_po_bank", "VB Catch-Weight", "2025-06-01")
    rid = run["id"]

    # Glen Rose chicken-style line: 10 cases shipped, 761 lbs total at $1.35/lb = $1,027.35.
    # Non-weight line: plain 12 cases @ $5.00 = $60.00.
    weights = [77.00, 76.00, 76.00, 78.00, 75.00, 73.00, 77.00, 78.00, 75.00, 76.00]
    assert abs(sum(weights) - 761.00) < 0.01
    result_store.save(rid, "vendor_bills", [{
        "id": "b-cw", "vendor_confirmed": "Glen Rose Meat Company",
        "vendor_extracted": "Glen Rose Meat Company",
        "invoice_number": "GR-001", "invoice_date": "2025-06-01",
        "extraction_status": "review",
        "lines": [
            # Catch-weight line — qty=10 cases, but Bill Import should use 761 lbs
            {"id": "l-cw", "bill_item_code": "0031110087-01",
             "description": "Chicken Wog 20HD 3.5 up (CVP)",
             "qty": 10, "rate": 1.35, "total": 1027.35,
             "individual_weights": weights,
             "needs_review": False, "user_confirmed": True},
            # Standard case-priced line — Bill Import qty stays as case count
            {"id": "l-std", "bill_item_code": "STD-1",
             "description": "Standard case item",
             "qty": 12, "rate": 5.0, "total": 60.0,
             "needs_review": False, "user_confirmed": True},
        ],
    }])
    result_store.save(rid, "po_bank", [])  # all lines will be "not_on_po"; that's fine

    res = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, "b-cw"))
    assert res["success"] is True, res

    rows = result_store.load(rid, "bill_import_rows", [])
    rows_by_code = {r["description"].split("\n")[0]: r for r in rows}

    cw_row = next(r for r in rows if "Chicken Wog" in (r.get("description") or ""))
    std_row = next(r for r in rows if "Standard case item" in (r.get("description") or ""))

    # PRD §5: weight-based qty must be total lbs, not the case count
    assert cw_row["qty"] == 761.0, (
        f"Catch-weight qty should be sum(individual_weights)=761.0 lbs, got {cw_row['qty']}"
    )
    # And price × qty must reconcile to the printed total (within rounding)
    assert abs(cw_row["price"] * cw_row["qty"] - cw_row["total"]) < 0.01, (
        f"price×qty ({cw_row['price']}×{cw_row['qty']}) must equal total ({cw_row['total']})"
    )

    # Standard line untouched — still case count
    assert std_row["qty"] == 12
    assert abs(std_row["price"] * std_row["qty"] - std_row["total"]) < 0.01

    run_store.delete_run(rid)


def test_vendor_bills_single_weight_lbs_line_uses_total_lbs_as_qty():
    """R.W. Zant-style line: one Qty (cases) + one LBS column (no per-piece list).
    AI prompt converts the single LBS number into individual_weights=[lbs_total]
    (see ai_extraction.py STEP 3c). Bill Import qty must come out as that lbs total
    so price × qty = total."""
    run = run_store.create_run("vendor_bill_po_bank", "VB Single LBS", "2025-06-01")
    rid = run["id"]
    # Qty=2 cases, LBS=80.00, Rate=$0.72, Amount=$57.60 — Zant-style row
    result_store.save(rid, "vendor_bills", [{
        "id": "b-sl", "vendor_confirmed": "R.W. Zant",
        "vendor_extracted": "R.W. Zant",
        "invoice_number": "R3407001", "invoice_date": "2025-06-01",
        "extraction_status": "review",
        "lines": [
            {"id": "l1", "bill_item_code": "Z-101",
             "description": "Beef cut catch-weight",
             "qty": 2, "rate": 0.72, "total": 57.60,
             "individual_weights": [80.00],
             "needs_review": False, "user_confirmed": True},
        ],
    }])
    result_store.save(rid, "po_bank", [])
    res = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, "b-sl"))
    assert res["success"] is True, res
    rows = result_store.load(rid, "bill_import_rows", [])
    assert len(rows) == 1
    assert rows[0]["qty"] == 80.0, f"Single-LBS qty must equal the total lbs (80.0), got {rows[0]['qty']}"
    assert abs(rows[0]["price"] * rows[0]["qty"] - rows[0]["total"]) < 0.01
    run_store.delete_run(rid)


def test_vendor_bills_safety_net_infers_weight_when_qty_rate_total_mismatch():
    """R.W. Zant bill safety net: if AI missed the '/LB' tag and left
    individual_weights null, but qty*rate != total and total/rate is clearly
    larger than the case count, derive qty = total/rate so the QB row
    reconciles. (line: THIGH MEAT 5 cases / 200 lbs / $2.20/lb / $440.00)"""
    run = run_store.create_run("vendor_bill_po_bank", "VB Safety Net", "2025-06-01")
    rid = run["id"]
    result_store.save(rid, "vendor_bills", [{
        "id": "b-sn", "vendor_confirmed": "R.W. Zant",
        "vendor_extracted": "R.W. Zant",
        "invoice_number": "R3407203", "invoice_date": "2025-06-01",
        "extraction_status": "review",
        "lines": [
            # No individual_weights — simulates AI missing the /LB marker
            {"id": "l1", "bill_item_code": "252738",
             "description": "THIGH MEAT BL/SL JMBO CVP",
             "qty": 5, "rate": 2.20, "total": 440.00,
             "needs_review": False, "user_confirmed": True},
            # Normal /CS line on the same bill — must NOT be touched
            {"id": "l2", "bill_item_code": "878676",
             "description": "BEEF SALISBURY STEAK FC 8",
             "qty": 1, "rate": 76.99, "total": 76.99,
             "needs_review": False, "user_confirmed": True},
        ],
    }])
    result_store.save(rid, "po_bank", [])
    res = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, "b-sn"))
    assert res["success"] is True
    rows = result_store.load(rid, "bill_import_rows", [])
    thigh = next(r for r in rows if "THIGH" in (r.get("description") or ""))
    salis = next(r for r in rows if "SALISBURY" in (r.get("description") or ""))
    assert thigh["qty"] == 200.0, f"Safety net should infer 200 lbs from total/rate, got {thigh['qty']}"
    assert abs(thigh["price"] * thigh["qty"] - thigh["total"]) < 0.01
    assert salis["qty"] == 1, "Per-CS line must keep its case count"
    run_store.delete_run(rid)


def test_vendor_bills_no_false_qty_mismatch_when_po_uses_lbs_unit():
    """R.W. Zant All Issues bug: bill row '2 cases / $2.20 / $440.00' was being
    compared against a PO row with quantity=200 (lbs) and emitting a fake
    'qty mismatch (-198)'. With unit-of-measure reconciliation, the matcher
    must infer that the PO is in lbs (total/rate = 200 ≈ po_qty) and report
    NO qty mismatch when 2 cases × 100 lbs each == 200 lbs ordered."""
    run = run_store.create_run("vendor_bill_po_bank", "VB UoM", "2025-06-01")
    rid = run["id"]
    result_store.save(rid, "vendor_bills", [{
        "id": "b-uom", "vendor_confirmed": "R.W. Zant",
        "vendor_extracted": "R.W. Zant",
        "invoice_number": "R3407203", "invoice_date": "2025-06-01",
        "extraction_status": "review",
        "lines": [
            # Zant line: 2 cases × $2.20/lb × 200 lbs total = $440.00
            # individual_weights null (worst case — AI missed /LB tag)
            {"id": "l1", "bill_item_code": "Z252738",
             "description": "CHICKEN THIGH MEAT BONELESS",
             "qty": 2, "rate": 2.20, "total": 440.00,
             "forced_po_id": "po-thigh",
             "needs_review": False, "user_confirmed": True},
            # And a TRUE qty mismatch (PO 5 cases, billed 3 cases, $5/case)
            {"id": "l2", "bill_item_code": "STD",
             "description": "STANDARD CASE",
             "qty": 3, "rate": 5.00, "total": 15.00,
             "forced_po_id": "po-std",
             "needs_review": False, "user_confirmed": True},
        ],
    }])
    result_store.save(rid, "po_bank", [
        # PO orders 200 LBS — different unit than bill's 2 CASES
        {"id": "po-thigh", "vendor": "R.W. Zant", "item_code": "Z252738",
         "description": "CHICKEN THIGH MEAT BONELESS", "quantity": 200, "po_cost": 2.20,
         "ref_number": "PO-1", "status": "unprocessed"},
        # PO orders 5 cases — same unit as bill, real shortage
        {"id": "po-std", "vendor": "R.W. Zant", "item_code": "STD",
         "description": "STANDARD CASE", "quantity": 5, "po_cost": 5.00,
         "ref_number": "PO-1", "status": "unprocessed"},
    ])
    res = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, "b-uom"))
    assert res["success"] is True
    blocks = result_store.load(rid, "vendor_summary_blocks", [])
    items = [it for b in blocks for it in b.get("items", [])]
    by_code = {it.get("item_code"): it for it in items}
    # The UoM-mismatched line must NOT carry a qty mismatch flag
    if "Z252738" in by_code:
        thigh = by_code["Z252738"]
        assert not thigh.get("has_qty_mismatch"), (
            f"Thigh meat row should not be flagged qty mismatch (UoM reconciled); got {thigh}"
        )
        assert thigh.get("section") != "qty_issue"
    # The genuine shortage must still appear as a qty mismatch
    assert "STD" in by_code, f"Standard line must reach summary_items; got codes {list(by_code)}"
    std = by_code["STD"]
    assert std.get("has_qty_mismatch"), f"Real qty mismatch must still be flagged; got {std}"
    run_store.delete_run(rid)


def test_vendor_bills_pricing_unit_drives_uom_reconciliation():
    """When AI returns pricing_unit per line (CS / LB / EA), the matcher must
    compare bill_qty in that unit against the PO row:
      - pricing_unit='LB' → compare lbs vs PO (which is in lbs)
      - pricing_unit='CS' → compare cases vs PO (which is in cases)
    A bill with one /LB row and one /CS row on the same vendor must produce
    no false qty mismatch on either, even though both rows print a weight column."""
    run = run_store.create_run("vendor_bill_po_bank", "VB PricingUnit", "2025-06-01")
    rid = run["id"]
    result_store.save(rid, "vendor_bills", [{
        "id": "b-pu", "vendor_confirmed": "R.W. Zant",
        "vendor_extracted": "R.W. Zant",
        "extraction_status": "review",
        "lines": [
            # /LB row — Zant THIGH MEAT (25 cs / 1000 lbs / $2.20/LB / $2200)
            {"id": "l-lb", "bill_item_code": "Z252738",
             "description": "THIGH MEAT BL/SL JMBO",
             "qty": 25, "rate": 2.20, "total": 2200.00,
             "pricing_unit": "LB",
             "forced_po_id": "po-thigh",
             "needs_review": False, "user_confirmed": True},
            # /CS row — Zant SOUR CREAM (10 cs / 320 lbs ship-weight / $40.20/CS / $402)
            # PO is in cases too; weight column 320 is shipping info only.
            {"id": "l-cs", "bill_item_code": "Z110997",
             "description": "SOUR CREAM TUB SUPER K",
             "qty": 10, "rate": 40.20, "total": 402.00,
             "pricing_unit": "CS",
             "forced_po_id": "po-sc",
             "needs_review": False, "user_confirmed": True},
        ],
    }])
    result_store.save(rid, "po_bank", [
        # PO in LBS — same as /LB bill row → no mismatch
        {"id": "po-thigh", "vendor": "R.W. Zant", "item_code": "Z252738",
         "description": "THIGH MEAT", "quantity": 1000, "po_cost": 2.20,
         "ref_number": "PO-1", "status": "unprocessed"},
        # PO in CASES — same as /CS bill row → no mismatch
        {"id": "po-sc", "vendor": "R.W. Zant", "item_code": "Z110997",
         "description": "SOUR CREAM", "quantity": 10, "po_cost": 40.20,
         "ref_number": "PO-1", "status": "unprocessed"},
    ])
    res = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, "b-pu"))
    assert res["success"] is True
    blocks = result_store.load(rid, "vendor_summary_blocks", [])
    items = {it["item_code"]: it for b in blocks for it in b.get("items", [])}
    # Both rows reach the summary (matched) but neither flags qty mismatch
    for code in ("Z252738", "Z110997"):
        if code in items:
            assert not items[code].get("has_qty_mismatch"), (
                f"{code} should not be flagged qty mismatch with pricing_unit driver; got {items[code]}"
            )
    run_store.delete_run(rid)


def test_vendor_bills_zero_shipment_line_is_highlighted_red():
    """OUT / unshipped lines (qty=0, total=$0) used to slip through unmarked
    because their math is technically valid. They must now carry
    highlight_status='zero_shipment' so the Bill Import sheet paints them red."""
    run = run_store.create_run("vendor_bill_po_bank", "VB Zero Shipment", "2025-06-01")
    rid = run["id"]
    result_store.save(rid, "vendor_bills", [{
        "id": "b-zs", "vendor_confirmed": "Acme", "vendor_extracted": "Acme",
        "extraction_status": "review",
        "lines": [
            # OUT / unshipped: vendor listed it but shipped none.
            {"id": "l1", "bill_item_code": "Z1", "description": "Backordered widget",
             "qty": 0, "rate": 0, "total": 0,
             "forced_po_id": "po-z1",
             "needs_review": False, "user_confirmed": True},
            # Normal line on same bill.
            {"id": "l2", "bill_item_code": "Z2", "description": "Shipped widget",
             "qty": 3, "rate": 2.00, "total": 6.00,
             "forced_po_id": "po-z2",
             "needs_review": False, "user_confirmed": True},
            # Unmatched-PO line, also highlighted but as not_on_po (separate case).
            {"id": "l3", "bill_item_code": "Z3", "description": "Mystery widget",
             "qty": 0, "rate": 0, "total": 0,
             "needs_review": False, "user_confirmed": True},
        ],
    }])
    result_store.save(rid, "po_bank", [
        {"id": "po-z1", "vendor": "Acme", "item_code": "Z1",
         "description": "Backordered widget", "quantity": 1, "po_cost": 0,
         "ref_number": "PO-1", "status": "unprocessed"},
        {"id": "po-z2", "vendor": "Acme", "item_code": "Z2",
         "description": "Shipped widget", "quantity": 3, "po_cost": 2.00,
         "ref_number": "PO-1", "status": "unprocessed"},
    ])
    res = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, "b-zs"))
    assert res["success"] is True
    rows = result_store.load(rid, "bill_import_rows", [])
    zs = next(r for r in rows if "Backordered" in (r.get("description") or ""))
    normal = next(r for r in rows if "Shipped widget" in (r.get("description") or ""))
    mystery = next(r for r in rows if "Mystery widget" in (r.get("description") or ""))
    # OUT line matched to a PO → zero_shipment red highlight
    assert zs["highlight_status"] == "zero_shipment", \
        f"Matched OUT row should be flagged zero_shipment, got {zs['highlight_status']!r}"
    # Normal shipped line — no highlight
    assert normal["highlight_status"] is None
    # Unmatched OUT line is already red as not_on_po (stronger flag wins)
    assert mystery["highlight_status"] == "not_on_po"
    run_store.delete_run(rid)


def test_vendor_bills_catch_weight_with_empty_weights_keeps_case_qty():
    """Safety: a line with individual_weights=[] (no actual weights) must NOT
    overwrite qty with 0 — fall back to the original case count."""
    run = run_store.create_run("vendor_bill_po_bank", "VB Empty Weights", "2025-06-01")
    rid = run["id"]
    result_store.save(rid, "vendor_bills", [{
        "id": "b-ew", "vendor_confirmed": "X", "vendor_extracted": "X",
        "extraction_status": "review",
        "lines": [
            {"id": "l1", "bill_item_code": "A", "description": "A",
             "qty": 5, "rate": 2.0, "total": 10.0,
             "individual_weights": [],
             "needs_review": False, "user_confirmed": True},
        ],
    }])
    result_store.save(rid, "po_bank", [])
    res = asyncio.run(vendor_bills_svc.process_confirmed_bill(rid, "b-ew"))
    assert res["success"] is True
    rows = result_store.load(rid, "bill_import_rows", [])
    assert rows[0]["qty"] == 5, "Empty individual_weights must not overwrite original qty"
    run_store.delete_run(rid)
