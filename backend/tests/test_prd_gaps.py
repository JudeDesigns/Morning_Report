"""Regression tests for the 10 PRD gap fixes."""
import os
import tempfile

STORAGE = tempfile.mkdtemp()
os.environ.setdefault("STORAGE_PATH", STORAGE)
os.environ.setdefault("SECRET_KEY", "prd-gap-tests-secret-key-xyz")

from decimal import Decimal


# ---------------------------------------------------------------------------
# Gap 1 — Jetro 'R' return indicator nets like void
# ---------------------------------------------------------------------------
def test_gap1_r_return_classification():
    from app.services.jetro.parser import normalize_code

    # Simulate parsed line behaviour by exercising fold_invoice_lines with a
    # return-typed line that has negative qty (parser converts R→negative).
    from app.services.jetro.reconciler import fold_invoice_lines

    base = {"base_code": "12345", "kind": "C", "full_code": "12345",
            "row_type": "item", "qty": Decimal("5"), "total": Decimal("50.00"),
            "description": "ITEM A", "_folded": []}
    ret = {"base_code": "12345", "kind": "C", "full_code": "12345",
           "row_type": "return", "qty": Decimal("-2"), "total": Decimal("-20.00"),
           "description": "ITEM A", "_folded": []}
    folded = fold_invoice_lines([dict(base), dict(ret)])
    assert len(folded) == 1
    assert folded[0]["qty"] == Decimal("3")
    assert folded[0]["total"] == Decimal("30.00")


# ---------------------------------------------------------------------------
# Gap 2 — finalize_unbilled_po does NOT run during single bill processing
# ---------------------------------------------------------------------------
def test_gap2_no_eager_po_not_charged():
    from app.services.vendor_bills import finalize_unbilled_po
    from app.store import runs as run_store, results as result_store

    run = run_store.create_run("vendor_bill_po_bank", "Gap2", "2025-01-01")
    rid = run["id"]
    # Two POs for the same vendor; one will be "matched" via a confirmed bill,
    # other should remain unprocessed until finalize is called.
    result_store.save(rid, "po_bank", [
        {"id": "po1", "vendor": "ACME", "item_code": "A1", "po_cost": 1.0,
         "quantity": 1, "status": "processed"},
        {"id": "po2", "vendor": "ACME", "item_code": "A2", "po_cost": 2.0,
         "quantity": 1, "status": "unprocessed"},
    ])
    # Bill marked confirmed for ACME so finalize will sweep leftovers
    result_store.save(rid, "vendor_bills", [
        {"id": "b1", "vendor_extracted": "ACME", "vendor_confirmed": "ACME",
         "extraction_status": "processed", "invoice_number": "INV-1", "lines": []},
    ])

    # Before finalize: po_not_charged should be empty
    pnc = result_store.load(rid, "po_not_charged", [])
    assert pnc == []

    result = finalize_unbilled_po(rid)
    assert result["added"] == 1
    pnc = result_store.load(rid, "po_not_charged", [])
    assert len(pnc) == 1
    assert pnc[0]["item_code"] == "A2"

    run_store.delete_run(rid)


def test_gap2_finalize_skips_vendors_without_confirmed_bill():
    from app.services.vendor_bills import finalize_unbilled_po
    from app.store import runs as run_store, results as result_store

    run = run_store.create_run("vendor_bill_po_bank", "Gap2b", "2025-01-01")
    rid = run["id"]
    result_store.save(rid, "po_bank", [
        {"id": "po1", "vendor": "OTHER", "item_code": "X1", "po_cost": 1.0,
         "quantity": 1, "status": "unprocessed"},
    ])
    result_store.save(rid, "vendor_bills", [])  # no confirmed bills

    result = finalize_unbilled_po(rid)
    assert result["added"] == 0
    assert result_store.load(rid, "po_not_charged", []) == []

    run_store.delete_run(rid)


# ---------------------------------------------------------------------------
# Gap 4 + 9 — integrity-mismatch blocks export, override unblocks
# ---------------------------------------------------------------------------
def test_gap4_and_9_override_records_and_unblocks():
    from app.store import runs as run_store, results as result_store

    run = run_store.create_run("jetro_reconciliation", "Gap4", "2025-01-01")
    rid = run["id"]
    result_store.save(rid, "jetro_invoices", [
        {"invoice_number": "I-1", "integrity_status": "mismatch",
         "printed_grand_total": 100, "processed_total": 99.50},
    ])
    assert not run_store.has_override(rid, "jetro_integrity_mismatch")
    run_store.record_override(
        rid, check="jetro_integrity_mismatch",
        reason="Verified manually with Restaurant Depot",
        user_id="user1", original_value="0.50",
    )
    updated = run_store.get_run(rid)
    assert len(updated["overrides"]) == 1
    assert updated["overrides"][0]["check"] == "jetro_integrity_mismatch"
    assert run_store.has_override(rid, "jetro_integrity_mismatch")

    run_store.delete_run(rid)


# ---------------------------------------------------------------------------
# Gap 7 — SP-remark below-cost yields a distinguishable reason
# ---------------------------------------------------------------------------
def test_gap7_sp_remark_distinguishable():
    from app.services.web_orders.checks import check_price_vs_cost

    sp_row = {"current_cost_price": Decimal("10"), "price": Decimal("8"),
              "remark": "SP for customer X"}
    plain_row = {"current_cost_price": Decimal("10"), "price": Decimal("8"),
                 "remark": "needs review"}

    sp_reasons = check_price_vs_cost(sp_row)
    plain_reasons = check_price_vs_cost(plain_row)
    assert sp_reasons and plain_reasons
    assert "SP override" in sp_reasons[0]
    assert "SP override" not in plain_reasons[0]
    # Token guard: 'spaghetti' should NOT be treated as SP
    spag = check_price_vs_cost({"current_cost_price": Decimal("10"),
                                 "price": Decimal("8"),
                                 "remark": "spaghetti sauce"})
    assert spag and "SP override" not in spag[0]


# ---------------------------------------------------------------------------
# Gap 10 — per-field needs_review flagging
# ---------------------------------------------------------------------------
def test_gap10_per_field_confidence_flagging():
    from app.services.ai_extraction import flag_low_confidence

    extracted = {
        "vendor": "ACME",
        "header_confidence": {"vendor": 0.99, "invoice_number": 0.5, "invoice_date": 0.95},
        "lines": [
            {"description": "A", "qty": 1, "rate": 1.0, "total": 1.0,
             "confidence": 0.95,
             "field_confidence": {"rate": 0.4, "total": 0.95}},
            {"description": "B", "qty": None, "rate": 2.0, "total": 4.0,
             "confidence": 0.95, "field_confidence": {}},
        ],
    }
    out = flag_low_confidence(extracted, threshold=0.9)
    assert "invoice_number" in out["header_needs_review"]
    assert out["lines"][0]["field_needs_review"] == ["rate"]
    assert out["lines"][0]["needs_review"] is True
    assert "qty" in out["lines"][1]["field_needs_review"]
    assert out["has_low_confidence"] is True



# ---------------------------------------------------------------------------
# Workflow v3 §10 — All Issues sheet builder
# ---------------------------------------------------------------------------
def _sample_inputs():
    """Synthetic inputs covering all three issue types per v3 §10."""
    bill_import_rows = [
        # Extra — billed not on PO
        {"highlight_status": "not_on_po", "qty": 2, "price": 5.5, "total": 11.0,
         "ref": "INV-7", "vendor": "Acme", "description": "PO desc\nVendor: Mushroom Med"},
        # Normal row — should NOT appear in All Issues
        {"highlight_status": None, "qty": 4, "price": 3.0, "total": 12.0,
         "ref": "INV-7", "vendor": "Acme", "description": "TOMATO 25LB"},
    ]
    summary_items_by_block = {
        "b1": [{
            "has_qty_mismatch": True, "section": "overcharged",
            "item_code": "M-001", "item_description": "Mushroom Med",
            "po_qty": 6, "bill_qty": 4, "qty_diff": -2, "invoice_rate": 7.0,
            "po_ref_number": "PO-101", "vendor": "Acme", "invoice_number": "INV-7",
        }],
        # A pure rate move with no qty diff — should NOT appear in All Issues
        "b2": [{
            "has_qty_mismatch": False, "section": "overcharged",
            "item_code": "O-001", "item_description": "Onion",
            "po_qty": 10, "bill_qty": 10, "qty_diff": 0, "invoice_rate": 6.0,
            "po_ref_number": "PO-101", "vendor": "Acme", "invoice_number": "INV-7",
        }],
    }
    po_not_charged = [
        {"item_code": "S-200", "description": "Springmix 4lb", "vendor": "Acme",
         "ref_number": "PO-101", "qty_ordered": 3, "po_cost": 12.5,
         "invoice_numbers": ["INV-7"]},
        # Larger dollar size Missing — must sort first within group
        {"item_code": "L-300", "description": "Lemon 40lb", "vendor": "Acme",
         "ref_number": "PO-102", "qty_ordered": 4, "po_cost": 25.0,
         "invoice_numbers": ["INV-7"]},
    ]
    return bill_import_rows, summary_items_by_block, [], po_not_charged


def test_v3_all_issues_row_types_and_filtering():
    """Missing/Extra/Qty mismatch are emitted; non-issues are excluded."""
    from app.services.vendor_bills.exporter import _build_all_issues_rows
    rows = _build_all_issues_rows(*_sample_inputs())
    types = [r["type"] for r in rows]
    assert types.count("Missing") == 2
    assert types.count("Extra") == 1
    assert types.count("Qty mismatch") == 1
    # Normal bill row and pure-rate item must not appear
    descriptions = [r["description"] for r in rows]
    assert "TOMATO 25LB" not in descriptions
    assert "Onion" not in descriptions


def test_v3_all_issues_sort_order():
    """Group order is Missing -> Extra -> Qty mismatch; within Missing, sort by
    dollar size desc (PO qty x cost)."""
    from app.services.vendor_bills.exporter import _build_all_issues_rows
    rows = _build_all_issues_rows(*_sample_inputs())
    types_seen = [r["type"] for r in rows]
    assert types_seen == ["Missing", "Missing", "Extra", "Qty mismatch"]
    # Lemon (4 * 25 = 100) sorts before Springmix (3 * 12.5 = 37.5)
    assert rows[0]["item_code"] == "L-300"
    assert rows[1]["item_code"] == "S-200"


def test_v3_all_issues_detail_templates():
    """Detail strings follow the v3 §10.3 templates."""
    from app.services.vendor_bills.exporter import _build_all_issues_rows
    rows = _build_all_issues_rows(*_sample_inputs())
    by_type = {r["type"]: r for r in rows if r["type"] != "Missing"}
    missing = [r for r in rows if r["type"] == "Missing"][1]  # smaller one
    assert "Ordered on PO PO-101 but not charged on invoice INV-7" in missing["detail"]
    assert "PO value ~$37.50" in missing["detail"]
    extra = by_type["Extra"]
    assert "Billed on invoice INV-7 but not on PO" in extra["detail"]
    assert "$11.00" in extra["detail"]
    qm = by_type["Qty mismatch"]
    assert "Invoice INV-7 billed 4; PO ordered 6 (-2)" in qm["detail"]
    # Line also moved on rate -> pointer appended
    assert "Cost Comparison" in qm["detail"]
    # Vendor-ref column carries vendor + invoice + PO
    assert qm["vendor_ref"] == "Acme \u00b7 Inv INV-7 \u00b7 PO PO-101"


def test_v3_all_issues_sheet_type_cell_colors():
    """build_all_issues_sheet shades the Type cell per v3 §10.3."""
    import openpyxl
    from app.services.vendor_bills.exporter import (
        _build_all_issues_rows, build_all_issues_sheet,
        ISSUE_MISSING_FILL, ISSUE_EXTRA_FILL, ISSUE_QTY_FILL,
    )
    rows = _build_all_issues_rows(*_sample_inputs())
    wb = openpyxl.Workbook()
    ws = wb.active
    build_all_issues_sheet(ws, rows)
    assert ws.cell(1, 1).value == "Type"
    expected_color = {
        "Missing": ISSUE_MISSING_FILL.fgColor.rgb,
        "Extra": ISSUE_EXTRA_FILL.fgColor.rgb,
        "Qty mismatch": ISSUE_QTY_FILL.fgColor.rgb,
    }
    for row_idx in range(2, ws.max_row + 1):
        t = ws.cell(row_idx, 1).value
        assert ws.cell(row_idx, 1).fill.fgColor.rgb == expected_color[t]


def test_v3_all_issues_sheet_position_in_workbook():
    """All Issues sits immediately after Summary Vendor & Warehouse (v3 §2)."""
    import types as _types
    from app.services.vendor_bills.exporter import generate_workbook
    import openpyxl
    run_ns = _types.SimpleNamespace(id="v3-sheet-order")
    out = generate_workbook(
        run=run_ns, po_rows=[], office_tasks=[], bill_import_rows=[],
        summary_blocks=[], summary_items_by_block={},
        po_not_charged=[], weight_rows=[], cost_comparison=[],
    )
    wb = openpyxl.load_workbook(out)
    names = wb.sheetnames
    assert names == [
        "Original PO - PO Bank",
        "Office and Driver's tasks",
        "Bill Import",
        "Summary Vendor & Warehouse",
        "All Issues",
        "PO Items Not Charged",
        "Individual Weights",
        "Cost Comparison",
    ]


def test_v3_po_bank_sheet_excludes_processed_rows():
    """Workflow v3 §3/§4.10/§17 — PO Bank retains only unprocessed rows."""
    import openpyxl
    from app.services.vendor_bills.exporter import build_po_bank_sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    po_rows = [
        {"ref_number": "PO-A", "item_code": "A1", "status": "unprocessed"},
        {"ref_number": "PO-B", "item_code": "B1", "status": "processed",
         "processed_bill_id": "bill-1"},
        {"ref_number": "PO-C", "item_code": "C1", "status": "unprocessed"},
    ]
    build_po_bank_sheet(ws, po_rows)
    refs = [ws.cell(r, 2).value for r in range(2, ws.max_row + 1)]
    assert refs == ["PO-A", "PO-C"]


def test_v3_compute_discrepancies_qty_fields():
    """Workflow v3 §10 requires per-line po_qty/bill_qty/qty_diff and a
    has_qty_mismatch flag that is independent of the rate-driven section."""
    from app.services.vendor_bills.matcher import compute_discrepancies
    bill = {"rate": 7.0, "qty": 4, "total": 28.0}
    po = {"po_cost": 5.0, "quantity": 6, "ref_number": "PO-101"}
    d = compute_discrepancies(bill, po, "Acme")
    assert d["section"] == "overcharged"  # rate moved -> rate-driven section
    assert d["has_qty_mismatch"] is True   # qty also differs
    assert d["po_qty"] == 6.0
    assert d["bill_qty"] == 4.0
    assert d["qty_diff"] == -2.0
    assert d["po_ref_number"] == "PO-101"


def test_v3_office_tasks_parser_layout():
    """Parser populates need_review/task per v3 §1.4: order_header -> vendor name
    + order note; item_task -> product description + task."""
    import openpyxl
    from app.services.vendor_bills.parser import parse_po_export
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["Terms", "RefNumber", "TxnDate", "Vendor", "Memo", "Total",
               "TxnLine Cost", "TxnLine Description", "TxnLine Item",
               "TxnLine Quantity", "Name", "CASE AVG WEIGHT", "UNIT AVG WEIGHT"])
    # Order header row (item code blank, cost 0)
    ws.append(["Net 30", "PO-1", None, "Acme", None, None, 0,
               "ORDER PLACED BY Barak", None, None, None, None, None])
    # Product row with embedded item task (blank line splits desc and task)
    ws.append(["Net 30", "PO-1", None, "Acme", None, None, 3.0,
               "GARLIC 5LB\n\nCheck rotation w/ driver", "G-001",
               10, "Produce", None, None])
    parsed = parse_po_export(wb)
    tasks = parsed["office_tasks"]
    header_task = next(t for t in tasks if t["task_type"] == "order_header")
    assert header_task["need_review"] == "Acme"
    assert header_task["task_instructions"] == "ORDER PLACED BY Barak"
    item_task = next(t for t in tasks if t["task_type"] == "item_task")
    assert item_task["need_review"] == "GARLIC 5LB"
    assert item_task["task_instructions"] == "Check rotation w/ driver"
