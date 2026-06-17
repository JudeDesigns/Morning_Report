"""Generate 5-sheet Jetro reconciliation workbook (PRD 9.8–9.12)."""
import uuid
from pathlib import Path
from decimal import Decimal
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from app.config import settings

HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
PINK_FILL = PatternFill("solid", fgColor="FFC7CE")
BOLD = Font(bold=True)


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_summary_sheet(
    ws,
    run,
    invoices: list[dict],
    issues: list[dict] | None = None,
    coupons: list[dict] | None = None,
    price_updates: list[dict] | None = None,
):
    """PRD §9.12 — at-a-glance Summary sheet."""
    issues = issues or []
    coupons = coupons or []
    price_updates = price_updates or []

    ws.title = "Summary"
    ws.append(["B&R Food Services — Jetro / Restaurant Depot Reconciliation"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["Run Name", run.name])
    ws.append(["Run Date", str(run.run_date)])
    ws.append([])

    ws.append(["Invoices"])
    ws.cell(ws.max_row, 1).font = BOLD
    for inv in invoices:
        ws.append(["Invoice #", inv.get("invoice_number"), "Type", inv.get("invoice_type"),
                   "Printed Total", inv.get("printed_grand_total"),
                   "Processed Total", inv.get("processed_total"),
                   "Integrity", inv.get("integrity_status")])
        # Highlight mismatches
        if inv.get("integrity_status") != "ok":
            for col in range(1, 11):
                ws.cell(ws.max_row, col).fill = PINK_FILL

    ws.append([])
    ws.append(["Totals"])
    ws.cell(ws.max_row, 1).font = BOLD

    total_invoices = len(invoices)
    credit_count = sum(1 for i in invoices if i.get("invoice_type") == "credit")
    grand_total_sum = sum((i.get("printed_grand_total") or 0) for i in invoices)
    integrity_ok = all(i.get("integrity_status") == "ok" for i in invoices) if invoices else True

    missing = sum(1 for i in issues if i.get("issue_type") == "Missing")
    extra = sum(1 for i in issues if i.get("issue_type") == "Extra")
    qty_mismatch = sum(1 for i in issues if i.get("issue_type") == "Qty mismatch")

    coupon_count = len(coupons)
    coupon_savings = sum((c.get("invoice_total_savings") or 0) for c in coupons)

    price_changes = len(price_updates)
    rises = sum(1 for p in price_updates
                if (p.get("cost_change_old_minus_new") or 0) > 0)
    drops = sum(1 for p in price_updates
                if (p.get("cost_change_old_minus_new") or 0) < 0)

    for label, value in [
        ("Total invoices", total_invoices),
        ("Credit invoices", credit_count),
        ("Sum of printed grand totals", grand_total_sum),
        ("All invoices integrity OK", "Yes" if integrity_ok else "No — see invoice rows"),
        ("Missing items (ordered, not billed)", missing),
        ("Extra items (billed, not ordered)", extra),
        ("Qty mismatches", qty_mismatch),
        ("Coupons captured", coupon_count),
        ("Total coupon savings (this invoice)", coupon_savings),
        ("Price changes detected", price_changes),
        ("  ↳ cost rises", rises),
        ("  ↳ cost drops", drops),
    ]:
        ws.append([label, value])
    # Format $ cells
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row_idx, 2)
        if isinstance(cell.value, (int, float)) and cell.value not in (total_invoices, credit_count,
                                                                       missing, extra, qty_mismatch,
                                                                       coupon_count, price_changes,
                                                                       rises, drops):
            cell.number_format = "#,##0.00"


def build_jetro_import_sheet(ws, import_rows: list[dict]):
    ws.title = "Jetro import"
    headers = ["Line", "UPC", "Item", "Description", "Price", "Total", "Qty", "Vendor", "Ref. #", "Date", "Type"]
    ws.append(headers)
    _style_header(ws)
    for row in import_rows:
        ws.append([
            row.get("line"), row.get("upc"), row.get("item_code"),
            row.get("description"), row.get("cost"), row.get("total"),
            row.get("qty"), row.get("vendor", "Jetro"),
            row.get("invoice_number"), row.get("invoice_date"),
            row.get("type", "Inventory Part"),
        ])
    # Format currency cols (Price=col5, Total=col6)
    for row_idx in range(2, ws.max_row + 1):
        for col in [5, 6]:
            ws.cell(row_idx, col).number_format = "#,##0.00"


def build_all_issues_sheet(ws, issues: list[dict]):
    ws.title = "All Issues"
    headers = ["Type", "Item Code", "Quantity", "Item Description", "Used by:", "Detail"]
    ws.append(headers)
    _style_header(ws)
    for issue in issues:
        ws.append([
            issue.get("issue_type"), issue.get("item_code"), issue.get("quantity"),
            issue.get("item_description"), issue.get("used_by"), issue.get("detail"),
        ])


def build_price_update_sheet(ws, price_updates: list[dict]):
    ws.title = "Price Update"
    headers = ["Item Code", "Qty Charged", "Item Description", "Old Cost", "New Cost", "Cost Change", "Used by:"]
    ws.append(headers)
    _style_header(ws)
    for update in price_updates:
        row_idx = ws.max_row + 1
        ws.append([
            update.get("item_code"), update.get("qty_charged"), update.get("item_description"),
            update.get("old_cost"), update.get("new_cost"), update.get("cost_change_old_minus_new"),
            update.get("used_by"),
        ])
        change = update.get("cost_change_old_minus_new")
        if change is not None:
            fill = GREEN_FILL if change < 0 else (PINK_FILL if change > 0 else None)
            if fill:
                for col in range(1, 8):
                    ws.cell(row_idx, col).fill = fill
    for row_idx in range(2, ws.max_row + 1):
        for col in [4, 5, 6]:
            ws.cell(row_idx, col).number_format = "#,##0.0000"


def build_coupons_sheet(ws, coupons: list[dict], invoice_number: str = ""):
    ws.title = "Coupons \u2013 Take Advantage"
    headers = [
        "Item Code", "Description", "Coupon Amount", "Qty",
        "Total Savings (invoice)", "Invoice #", "8 Weeks Usage", "8wk Usage × Coupon $",
    ]
    ws.append(headers)
    _style_header(ws)
    for coupon in coupons:
        ws.append([
            coupon.get("item_code"), coupon.get("description"), coupon.get("coupon_amount"),
            coupon.get("qty"), coupon.get("invoice_total_savings"), invoice_number,
            coupon.get("eight_week_usage"), coupon.get("projected_savings"),
        ])
    for row_idx in range(2, ws.max_row + 1):
        for col in [3, 5, 7, 8]:
            ws.cell(row_idx, col).number_format = "#,##0.00"


def generate_workbook(
    run,
    invoices: list[dict],
    import_rows: list[dict],
    issues: list[dict],
    price_updates: list[dict],
    coupons: list[dict],
) -> str:
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Summary")
    ws_import = wb.create_sheet("Jetro import")
    ws_issues = wb.create_sheet("All Issues")
    ws_price = wb.create_sheet("Price Update")
    ws_coupons = wb.create_sheet("Coupons \u2013 Take Advantage")

    build_summary_sheet(ws_summary, run, invoices, issues, coupons, price_updates)
    build_jetro_import_sheet(ws_import, import_rows)
    build_all_issues_sheet(ws_issues, issues)
    build_price_update_sheet(ws_price, price_updates)
    invoice_num = invoices[0]["invoice_number"] if invoices else ""
    build_coupons_sheet(ws_coupons, coupons, invoice_num)

    storage_root = Path(settings.STORAGE_PATH) / str(run.id) / "exports"
    storage_root.mkdir(parents=True, exist_ok=True)
    out_path = storage_root / f"Jetro_Reconciliation_{run.id}.xlsx"
    wb.save(str(out_path))
    return str(out_path)
