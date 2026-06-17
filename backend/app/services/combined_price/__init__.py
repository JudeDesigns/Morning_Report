"""Combined Price Changes service (PRD Module D) — file-based, no DB."""
from decimal import Decimal
from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment

from app.config import settings
from app.store import results as result_store
from app.store.runs import get_run, update_run
from app.services.vendor_bills.exporter import _build_all_issues_rows

NO_CHANGE_THRESHOLD = Decimal(str(settings.PRICE_CHANGE_NO_CHANGE_THRESHOLD))
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor="C6EFCE")
PINK_FILL = PatternFill("solid", fgColor="FFC7CE")

# All Issues type-cell fills (match vendor_bills exporter v3 §10.3)
ISSUE_MISSING_FILL = PatternFill("solid", fgColor="DDEBF7")
ISSUE_EXTRA_FILL = PatternFill("solid", fgColor="FCE4D6")
ISSUE_QTY_FILL = PatternFill("solid", fgColor="FFF2CC")


async def process_run(
    run_id: str,
    jetro_run_id: str | None = None,
    vendor_run_id: str | None = None,
) -> dict:
    """Pull price changes from Jetro and Vendor Bill JSON results → combined output."""
    rows = []

    # ── Jetro Price Updates (Old-Cost → New-Cost, diff = New − Old) ───────────
    if jetro_run_id:
        price_updates = result_store.load(jetro_run_id, "jetro_price_updates", [])
        for pu in price_updates:
            old = Decimal(str(pu["old_cost"])) if pu.get("old_cost") is not None else None
            new = Decimal(str(pu["new_cost"])) if pu.get("new_cost") is not None else None
            if old is None or new is None:
                continue
            diff = new - old
            if abs(diff) <= NO_CHANGE_THRESHOLD:
                continue
            rows.append({
                "item_code": pu.get("item_code"),
                "item_description": pu.get("item_description"),
                "old_cost": float(old),
                "new_cost": float(new),
                "difference": float(diff),
                "used_by": pu.get("used_by"),
                "source": "Jetro / Restaurant Depot",
            })

    # ── Vendor Bill Cost Comparison (diff already New − Old convention) ───────
    if vendor_run_id:
        cost_comparison = result_store.load(vendor_run_id, "cost_comparison", [])
        for cc in cost_comparison:
            diff_raw = cc.get("difference")
            if diff_raw is None:
                continue
            diff = Decimal(str(diff_raw))
            if abs(diff) <= NO_CHANGE_THRESHOLD:
                continue
            rows.append({
                "item_code": cc.get("item_code"),
                "item_description": cc.get("description"),
                "old_cost": cc.get("po_cost"),
                "new_cost": cc.get("vendor_bill_cost"),
                "difference": float(diff),
                "used_by": None,
                "source": f"Vendor bills ({cc['vendor']})" if cc.get("vendor") else "Vendor bills",
            })

    # Sort: reds (price DECREASES, diff < 0) on top, then greens (INCREASES, diff > 0),
    # zero/None at the bottom. Within each colour group, largest |diff| first so the
    # biggest moves appear first inside the section.
    def _sort_key(r):
        d = r.get("difference")
        if d is None or d == 0:
            return (2, 0.0)            # neutral last
        if d < 0:
            return (0, -abs(d))        # reds first, biggest |drop| on top
        return (1, -abs(d))            # greens after, biggest rise on top
    rows.sort(key=_sort_key)

    # ── Consolidated All Issues (Jetro + Vendor) ──────────────────────────────
    issues = _build_combined_issues_rows(jetro_run_id, vendor_run_id)

    result_store.save(run_id, "combined_price_rows", rows)
    result_store.save(run_id, "combined_issues_rows", issues)
    update_run(run_id, {"status": "processed"})

    return {
        "success": True,
        "total_changes": len(rows),
        "increases": sum(1 for r in rows if (r.get("difference") or 0) > 0),
        "decreases": sum(1 for r in rows if (r.get("difference") or 0) < 0),
        "total_issues": len(issues),
    }


def _build_combined_issues_rows(
    jetro_run_id: str | None,
    vendor_run_id: str | None,
) -> list[dict]:
    """Normalise issues from both sources into a single shape for the All Issues sheet."""
    rows: list[dict] = []

    if jetro_run_id:
        for it in result_store.load(jetro_run_id, "jetro_issues", []):
            rows.append({
                "source": "Jetro / Restaurant Depot",
                "type": it.get("issue_type") or "",
                "item_code": it.get("item_code") or "",
                "quantity": it.get("quantity"),
                "description": it.get("item_description") or "",
                "reference": it.get("used_by") or "",
                "detail": it.get("detail") or "",
                "_dollar_size": float(it.get("dollar_size") or 0),
            })

    if vendor_run_id:
        bill_import_rows = result_store.load(vendor_run_id, "bill_import_rows", [])
        summary_blocks = result_store.load(vendor_run_id, "vendor_summary_blocks", [])
        items_by_block = {b["id"]: b.get("items", []) for b in summary_blocks}
        po_not_charged = result_store.load(vendor_run_id, "po_not_charged", [])
        vendor_issues = _build_all_issues_rows(
            bill_import_rows, items_by_block, summary_blocks, po_not_charged,
        )
        for it in vendor_issues:
            rows.append({
                "source": "Vendor bills",
                "type": it.get("type") or "",
                "item_code": it.get("item_code") or "",
                "quantity": it.get("qty_ordered"),
                "description": it.get("description") or "",
                "reference": it.get("vendor_ref") or "",
                "detail": it.get("detail") or "",
                "_dollar_size": float(it.get("_dollar_size") or 0),
            })

    type_order = {"Missing": 0, "Extra": 1, "Qty mismatch": 2}
    source_order = {"Jetro / Restaurant Depot": 0, "Vendor bills": 1}
    rows.sort(key=lambda r: (
        type_order.get(r["type"], 99),
        source_order.get(r["source"], 99),
        -r["_dollar_size"],
    ))
    return rows


async def export_workbook(run_id: str) -> str:
    rows = result_store.load(run_id, "combined_price_rows", [])
    issues = result_store.load(run_id, "combined_issues_rows", [])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Price changes from both sources"

    headers = ["Item Code", "Item Description", "Old Cost", "New Cost", "Difference", "Used by:", "Source"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for row in rows:
        row_idx = ws.max_row + 1
        ws.append([row.get("item_code"), row.get("item_description"),
                   row.get("old_cost"), row.get("new_cost"),
                   row.get("difference"), row.get("used_by"), row.get("source")])
        diff = row.get("difference")
        if diff is not None:
            fill = GREEN_FILL if diff < 0 else (PINK_FILL if diff > 0 else None)
            if fill:
                for col in range(1, 8):
                    ws.cell(row_idx, col).fill = fill
        for col in [3, 4, 5]:
            ws.cell(row_idx, col).number_format = "#,##0.0000"

    # ── All Issues sheet (Jetro + Vendor consolidated) ────────────────────────
    _build_all_issues_sheet(wb.create_sheet("All Issues"), issues)

    storage_root = Path(settings.STORAGE_PATH) / run_id / "exports"
    storage_root.mkdir(parents=True, exist_ok=True)
    out_path = storage_root / f"Price_Changes_{run_id}.xlsx"
    wb.save(str(out_path))
    return str(out_path)


def _build_all_issues_sheet(ws, issues: list[dict]):
    """Consolidated cross-source All Issues sheet (Jetro + Vendor bills)."""
    headers = ["Source", "Type", "Item Code", "Quantity",
               "Item Description", "Reference (Customer / Vendor · Invoice · PO)", "Detail"]
    ws.append(headers)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for r in issues:
        row_idx = ws.max_row + 1
        ws.append([
            r.get("source"), r.get("type"), r.get("item_code"),
            r.get("quantity"), r.get("description"),
            r.get("reference"), r.get("detail"),
        ])
        t = r.get("type")
        fill = (ISSUE_MISSING_FILL if t == "Missing"
                else ISSUE_EXTRA_FILL if t == "Extra"
                else ISSUE_QTY_FILL if t == "Qty mismatch"
                else None)
        if fill:
            ws.cell(row_idx, 2).fill = fill
