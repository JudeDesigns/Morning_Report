"""Generate 4-sheet Excel workbook for Web Orders Check (PRD 8.8)."""
import io
import uuid
from pathlib import Path
from typing import List
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from app.config import settings

YELLOW_FILL = PatternFill("solid", fgColor="FFFF00")
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF")
PROBLEM_FILL = PatternFill("solid", fgColor="FFD9D9")

# PRD §8.4 — preserve historical Excel precision on enriched columns
# U=21, V=22, Y=25, Z=26, AB=28, AC=29, AD=30
NUM_FORMAT_COLS = (21, 22, 25, 26, 28, 29, 30)
NUM_FORMAT = "0.####"


def _header_style(ws, row=1):
    for cell in ws[row]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True)


def _auto_filter(ws):
    ws.auto_filter.ref = ws.dimensions


def _freeze_top_row(ws):
    ws.freeze_panes = "A2"


def _coerce_number(v):
    """Return int for whole-number floats so Excel never shows a trailing decimal point."""
    if isinstance(v, float) and v == int(v) and not (v != v):  # exclude NaN
        return int(v)
    return v


def _apply_number_formats(ws):
    """Apply 0.#### (or 0 for integers) format to PRD §8.4 columns U,V,Y,Z,AB,AC,AD."""
    for row_idx in range(2, ws.max_row + 1):
        for col in NUM_FORMAT_COLS:
            cell = ws.cell(row_idx, col)
            if cell.value is None or cell.value == "":
                continue
            # Coerce whole-number floats to int so Excel renders them cleanly
            cell.value = _coerce_number(cell.value)
            cell.number_format = "0" if isinstance(cell.value, int) else NUM_FORMAT


# Mapped columns owned by the parser (A,B,G,H,I,J,L,O,Q,S — 1-indexed).
# Passthrough columns: C,D,E,F,K,M,N,P,R (1-indexed indices 3,4,5,6,11,13,14,16,18).
# Underlying parser passthrough dict keys are 0-indexed: 2,3,4,5,10,12,13,15,17.
PASSTHROUGH_COL_INDICES = (2, 3, 4, 5, 10, 12, 13, 15, 17)

# Enrichment headers (T..AF) per PRD §8.3 — fixed labels.
ENRICHED_HEADERS = [
    "Category Name", "Current Cost Price", "Current Selling Price",          # T,U,V
    "Product", "New Bin", "Unit Price", "Case Price",                         # W,X,Y,Z
    "Qty On Hand", "Cost", "Case Avg Weight", "Unit Avg Weight",              # AA,AB,AC,AD
    "BIN Internal", "Unit",                                                   # AE,AF
]

# Fallback labels when no source-header is supplied (legacy generic names).
DEFAULT_PASSTHROUGH_LABELS = {
    2: "Col C", 3: "Col D", 4: "Col E", 5: "Col F", 10: "Col K",
    12: "Col M", 13: "Col N", 15: "Col P", 17: "Col R",
}


def _build_all_orders_headers(source_headers: List) -> List[str]:
    """Compose the 32-column All Orders header row.

    Uses real source headers for the passthrough columns when available so the
    output workbook preserves what the source spreadsheet called them.
    """
    def src(idx: int, default: str) -> str:
        if source_headers and idx < len(source_headers):
            val = source_headers[idx]
            if val:
                return str(val).strip()
        return default

    headers = [
        src(0, "Product Name"),                              # A
        src(1, "Code"),                                      # B
        src(2, DEFAULT_PASSTHROUGH_LABELS[2]),               # C
        src(3, DEFAULT_PASSTHROUGH_LABELS[3]),               # D
        src(4, DEFAULT_PASSTHROUGH_LABELS[4]),               # E
        src(5, DEFAULT_PASSTHROUGH_LABELS[5]),               # F
        src(6, "Qty"),                                       # G
        src(7, "Weight"),                                    # H
        src(8, "Individual Weight/Status"),                  # I
        src(9, "Price/Selling"),                             # J
        src(10, DEFAULT_PASSTHROUGH_LABELS[10]),             # K
        src(11, "Customer Name"),                            # L
        src(12, DEFAULT_PASSTHROUGH_LABELS[12]),             # M
        src(13, DEFAULT_PASSTHROUGH_LABELS[13]),             # N
        src(14, "Transaction Date"),                         # O
        src(15, DEFAULT_PASSTHROUGH_LABELS[15]),             # P
        src(16, "Route"),                                    # Q
        src(17, DEFAULT_PASSTHROUGH_LABELS[17]),             # R
        src(18, "Remark"),                                   # S
    ]
    headers.extend(ENRICHED_HEADERS)                          # T..AF
    return headers


def _build_all_orders_row(row: dict) -> list:
    """Build the 32-cell data row matching the header layout."""
    pt = row.get("passthrough") or {}
    n = _coerce_number  # shorthand

    # Normalize JSON-deserialised passthrough keys (may be str after JSON round-trip)
    def p(i: int):
        return n(pt.get(i, pt.get(str(i))))

    return [
        row.get("product_name"), row.get("code"),
        p(2), p(3), p(4), p(5),                              # C–F passthrough
        n(row.get("qty")), n(row.get("weight")), row.get("individual_weight_status"),
        n(row.get("price")),
        p(10),                                               # K passthrough
        row.get("customer_name"),
        p(12), p(13),                                        # M, N passthrough
        row.get("transaction_date"),
        p(15),                                               # P passthrough
        row.get("route"),
        p(17),                                               # R passthrough
        row.get("remark"),
        row.get("category_name"), n(row.get("current_cost_price")), n(row.get("current_selling_price")),
        n(row.get("shopping_product")), n(row.get("new_bin")), n(row.get("unit_price")), n(row.get("case_price")),
        n(row.get("quantity_on_hand")), n(row.get("inventory_cost")), n(row.get("case_avg_weight")),
        n(row.get("unit_avg_weight")), row.get("bin_internal"), row.get("unit"),
    ]


def build_all_orders_sheet(ws, rows: List[dict], source_headers: List = ()):
    ws.title = "All Orders"
    headers = _build_all_orders_headers(list(source_headers))
    ws.append(headers)
    _header_style(ws)
    _freeze_top_row(ws)
    _auto_filter(ws)

    for row in rows:
        if row.get("_empty") or row.get("is_spacer"):
            ws.append([""] * len(headers))
            continue
        ws.append(_build_all_orders_row(row))
    _apply_number_formats(ws)


def build_future_orders_sheet(ws, rows: List[dict], source_headers: List = ()):
    ws.title = "Future Orders"
    headers = _build_all_orders_headers(list(source_headers))
    ws.append(headers)
    _header_style(ws)
    _freeze_top_row(ws)
    _auto_filter(ws)
    for row in rows:
        if row.get("_empty") or row.get("is_spacer"):
            ws.append([""] * len(headers))
            continue
        ws.append(_build_all_orders_row(row))
    _apply_number_formats(ws)


PROBLEMATIC_HEADERS = [
    "Customer Name", "Problem", "Qty Ordered", "Weight", "Individual Weight",
    "Remark", "Product Name",
]


def build_problematic_items_sheet(ws, rows: List[dict], source_headers: List = ()):
    ws.title = "Problematic Items"
    all_orders_headers = _build_all_orders_headers(list(source_headers))
    ws.append(PROBLEMATIC_HEADERS + all_orders_headers)
    _header_style(ws)
    _freeze_top_row(ws)
    _auto_filter(ws)

    for row in rows:
        problems = row.get("problem_reasons") or []
        problem_text = "; ".join(problems) if problems else ""
        prefix = [
            row.get("customer_name"), problem_text, row.get("qty"), row.get("weight"),
            row.get("individual_weight_status"), row.get("remark"), row.get("product_name"),
        ]
        ws.append(prefix + _build_all_orders_row(row))
        # Highlight problem column (B = index 2)
        ws.cell(row=ws.max_row, column=2).fill = PROBLEM_FILL


BY_CUSTOMER_HEADERS = [
    "Customer Name", "Qty Ordered", "Product Name", "Code", "Notes",
    "Price", "Weight", "Status", "Current Selling Price", "Unit Avg Weight", "Case Avg Weight",
]


def build_by_customer_sheet(ws, rows: List[dict]):
    ws.title = "By Customer Problematic"
    ws.append(BY_CUSTOMER_HEADERS)
    _header_style(ws)
    _freeze_top_row(ws)
    _auto_filter(ws)

    n = _coerce_number
    sorted_rows = sorted(rows, key=lambda r: (r.get("customer_name") or ""))
    for row in sorted_rows:
        ws.append([
            row.get("customer_name"), n(row.get("qty")), row.get("product_name"),
            row.get("code"), row.get("remark"), n(row.get("price")), n(row.get("weight")),
            row.get("individual_weight_status"), n(row.get("current_selling_price")),
            n(row.get("unit_avg_weight")), n(row.get("case_avg_weight")),
        ])


def generate_workbook(
    same_day_rows: List[dict],
    future_rows: List[dict],
    problematic_rows: List[dict],
    run_id: uuid.UUID,
    source_headers: List = (),
) -> str:
    wb = openpyxl.Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    ws_all = wb.create_sheet("All Orders")
    ws_future = wb.create_sheet("Future Orders")
    ws_prob = wb.create_sheet("Problematic Items")
    ws_cust = wb.create_sheet("By Customer Problematic")

    build_all_orders_sheet(ws_all, same_day_rows, source_headers)
    build_future_orders_sheet(ws_future, future_rows, source_headers)
    build_problematic_items_sheet(ws_prob, problematic_rows, source_headers)
    build_by_customer_sheet(ws_cust, problematic_rows)

    # Save to storage
    storage_root = Path(settings.STORAGE_PATH) / str(run_id) / "exports"
    storage_root.mkdir(parents=True, exist_ok=True)
    out_path = storage_root / f"Web_Orders_Check_{run_id}.xlsx"
    wb.save(str(out_path))
    return str(out_path)
