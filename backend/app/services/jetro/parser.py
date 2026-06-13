"""Parse Jetro source sheet and Restaurant Depot invoices (PRD 9.3–9.6)."""
from decimal import Decimal, InvalidOperation
from typing import Optional
import openpyxl


def normalize_code(val) -> Optional[str]:
    if val is None:
        return None
    s = str(val).strip()
    if not s:
        return None
    if s.endswith(".0") and s[:-2].isdigit():
        s = s[:-2]
    return s


def safe_decimal(val) -> Optional[Decimal]:
    if val is None:
        return None
    try:
        return Decimal(str(val).strip())
    except (InvalidOperation, ValueError):
        return None


def parse_code_kind(raw_code: str) -> tuple[str, str]:
    """Return (base_code, kind) where kind is 'U' or 'C'."""
    if not raw_code:
        return ("", "C")
    if raw_code.upper().endswith("U"):
        return (raw_code[:-1], "U")
    return (raw_code, "C")


INTERNAL_WAREHOUSE_NAME = "Warehouse INVENTORY Order"


def parse_jetro_source(wb: openpyxl.Workbook) -> list[dict]:
    """Parse Jetro source sheet. PRD 9.3."""
    ws = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "jetro" in nl or "source" in nl or "stage" in nl or "sheet" in nl:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    rows = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        if not header_found:
            header_found = True
            continue  # skip header

        raw = list(row)

        def get(idx):
            return raw[idx] if idx < len(raw) else None

        # PRD cols: C=qty(2), D=product(3), E=code(4), O=customer(14),
        #           X=cost(23), Y=selling(24), AF=case_avg(31), AI=unit(34)
        raw_code = normalize_code(get(4))
        base_code, kind = parse_code_kind(raw_code or "")
        customer = str(get(14)).strip() if get(14) else None
        is_internal = customer == INTERNAL_WAREHOUSE_NAME if customer else False

        rows.append({
            "qty": safe_decimal(get(2)),
            "product_name": str(get(3)).strip() if get(3) else None,
            "raw_code": raw_code,
            "base_code": base_code,
            "kind": kind,
            "full_code": raw_code,
            "customer_name": customer,
            "current_cost_price": safe_decimal(get(23)),
            "current_selling_price": safe_decimal(get(24)),
            "case_avg_weight": safe_decimal(get(31)),
            "unit": str(get(34)).strip() if get(34) else None,
            "is_internal_inventory": is_internal,
            "_raw": raw,
        })
    return rows


def parse_rd_invoice_xlsx(wb: openpyxl.Workbook) -> dict:
    """Parse Restaurant Depot .xlsx invoice. PRD 9.4.
    Header at row 17, data from row 18 until first Sub-Total row."""
    ws = None
    for name in wb.sheetnames:
        nl = name.lower()
        if "invoice" in nl or "sheet" in nl or len(wb.sheetnames) == 1:
            ws = wb[name]
            break
    if ws is None:
        ws = wb.active

    invoice_number = None
    invoice_date = None
    grand_total = None
    lines = []

    rows_iter = ws.iter_rows(values_only=True)
    all_rows = list(rows_iter)

    # Scan preamble for Invoice number and date
    for i, row in enumerate(all_rows[:20]):
        row_str = " ".join(str(c) for c in row if c)
        if invoice_number is None and "Invoice" in row_str and "Convert" not in row_str:
            # PRD §9.4 — find the cell labelled "Invoice" / "Invoice #" /
            # "Invoice No" and read the first non-empty numeric/alphanumeric
            # value to its right. Avoid false hits on date or PO cells.
            cells = list(row)
            invoice_label_idx = None
            for j, c in enumerate(cells):
                if c is None:
                    continue
                s = str(c).strip().lower()
                if s.startswith("invoice"):
                    # Inline form like "Invoice # 12345"
                    rest = str(c).strip()
                    # Pull trailing digit run
                    digits = "".join(ch for ch in rest if ch.isdigit())
                    if digits and len(digits) >= 5:  # invoice numbers are typically 5+ digits
                        invoice_number = digits
                        break
                    invoice_label_idx = j
                    break
            if invoice_number is None and invoice_label_idx is not None:
                # Read cells to the right of the label
                for c in cells[invoice_label_idx + 1:]:
                    if c is None or str(c).strip() == "":
                        continue
                    s = str(c).strip()
                    # Skip obvious date strings
                    if "/" in s or "-" in s:
                        continue
                    digits = "".join(ch for ch in s if ch.isdigit())
                    if digits and len(digits) >= 5:
                        invoice_number = digits
                        break
        if "Date" in row_str:
            for cell in row:
                if cell and hasattr(cell, "year"):
                    invoice_date = cell
                elif cell and "/" in str(cell):
                    try:
                        from datetime import datetime
                        invoice_date = datetime.strptime(str(cell).strip(), "%m/%d/%Y").date()
                    except Exception:
                        pass
        if "Grand Total" in row_str or "grand total" in row_str.lower():
            for cell in reversed(row):
                d = safe_decimal(cell)
                if d is not None:
                    grand_total = d
                    break

    # Find header row (row 17 = index 16)
    data_start = 17  # 0-indexed
    for i, row in enumerate(all_rows):
        row_str = " ".join(str(c).lower() for c in row if c)
        if "description" in row_str and "price" in row_str:
            data_start = i + 1
            break

    # Parse data rows
    for row in all_rows[data_start:]:
        raw = list(row)
        if not any(raw):
            continue

        def get(idx):
            return raw[idx] if idx < len(raw) else None

        item_raw = str(get(2)).strip() if get(2) else ""
        desc = str(get(3)).strip() if get(3) else ""
        upc = str(get(1)).strip() if get(1) else ""

        # Terminator rows — Restaurant Depot puts "Sub-Total" / "Grand Total"
        # in column A (index 0); the test fixture puts "Grand Total" in column D.
        # Scan every cell so we catch both forms.
        joined_lower = " ".join(str(c).lower() for c in raw if c)
        if "grand total" in joined_lower:
            t = safe_decimal(get(7))
            if t is not None:
                grand_total = t
            break
        if "sub-total" in joined_lower or "subtotal" in joined_lower:
            # Continue scanning past Sub-Total so we can still pick up the
            # Grand Total row a few lines below, but skip this row's data.
            continue

        # Drop banner/footer rows that carry no item content (e.g. "Delivery
        # Charge" or stray label cells in column A). A real item, surcharge,
        # coupon, void, or return always has at least one of item_raw / upc /
        # description plus a numeric column.
        if not item_raw and not upc and not desc:
            continue

        # Detect row type
        is_surcharge = item_raw.lower() == "surcharge" or "crv" in desc.lower()
        is_coupon = upc.lower() == "coupon" or (safe_decimal(get(4)) is not None and safe_decimal(get(4)) < 0)
        qty_val = get(6)
        qty_upper = str(qty_val).upper() if qty_val else ""
        is_void = "V" in qty_upper
        is_return = (not is_void) and "R" in qty_upper

        qty_clean = safe_decimal(str(qty_val).replace("V", "").replace("R", "").strip()) if qty_val else None
        # Voids/returns net against the matching item — store qty as negative
        if (is_void or is_return) and qty_clean is not None and qty_clean > 0:
            qty_clean = -qty_clean
        cu = str(get(5)).strip() if get(5) else ""
        base_code = normalize_code(item_raw) if item_raw.isdigit() or (item_raw and item_raw.replace(".", "").isdigit()) else None
        kind = "U" if cu.upper().startswith("U") else "C"
        full_code = (base_code + "U") if base_code and kind == "U" else base_code

        row_type = "item"
        if is_surcharge:
            row_type = "surcharge"
        elif is_coupon:
            row_type = "coupon"
        elif is_void:
            row_type = "void"
        elif is_return:
            row_type = "return"

        lines.append({
            "line_number": str(get(0)) if get(0) else None,
            "upc": upc,
            "item_raw": item_raw,
            "description": desc,
            "price": safe_decimal(get(4)),
            "cu": cu,
            "qty": qty_clean,
            "total": safe_decimal(get(7)),
            "base_code": base_code,
            "kind": kind,
            "full_code": full_code,
            "row_type": row_type,
            "_raw": raw,
        })

    return {
        "invoice_number": invoice_number,
        "invoice_date": invoice_date,
        "grand_total": grand_total,
        "lines": lines,
    }


def parse_sales_per_week(wb: openpyxl.Workbook) -> dict:
    """Parse Sales-Per-Week report. Returns dict keyed by item code -> weekly_sales.

    Column A is the item code. The weekly sales value is located by header
    match ("Sales/Week"-style label); if no header matches, fall back to the
    rightmost column whose data is numeric. This tolerates layouts that
    interleave a Description column between the code and the sales number.
    """
    ws = wb.active
    sales_col = None
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        if not header_found:
            for idx, cell in enumerate(row):
                if cell is None:
                    continue
                label = str(cell).strip().lower().replace(" ", "")
                if "sales/week" in label or "salesperweek" in label or "sales_week" in label:
                    sales_col = idx
                    break
            header_found = True
            continue
        if sales_col is None:
            for idx in range(len(row) - 1, 0, -1):
                if safe_decimal(row[idx]) is not None:
                    sales_col = idx
                    break
        break

    if sales_col is None:
        sales_col = 1

    result = {}
    header_seen = False
    for row in ws.iter_rows(values_only=True):
        if not any(row):
            continue
        if not header_seen:
            header_seen = True
            continue
        raw = list(row)
        code = normalize_code(raw[0] if raw else None)
        if not code:
            continue
        sales = safe_decimal(raw[sales_col] if len(raw) > sales_col else None)
        if sales is not None:
            result[code] = sales
    return result
