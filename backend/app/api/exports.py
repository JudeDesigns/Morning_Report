import io
import zipfile
from pathlib import Path
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import FileResponse as FastAPIFileResponse, StreamingResponse
from app.core.auth import get_current_user, require_roles
from app.store import runs as run_store, results as result_store, audit as audit_store
from app.services import web_orders as wo_service
from app.services import jetro as jetro_service
from app.services import vendor_bills as vb_service
from app.services import combined_price as cp_service

router = APIRouter(prefix="/exports", tags=["exports"])

XLSX_MEDIA = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
_WORKFLOW_TYPES = {
    "web_orders_check": "web orders",
    "jetro_reconciliation": "Jetro reconciliation",
    "vendor_bill_po_bank": "vendor bills",
    "combined_price_changes": "combined price",
}
# Roles that may export per PRD §17.2
_EXPORT_ROLES = ("admin", "accounting", "office", "management")


def _resolve(run_id: str, expected_workflow: str, current_user: dict) -> dict:
    run = run_store.get_run(run_id)
    if not run:
        raise HTTPException(404, "Run not found")
    if run.get("workflow_type") != expected_workflow:
        raise HTTPException(400, f"Run is not a {_WORKFLOW_TYPES.get(expected_workflow)} run")
    if current_user.get("role") != "admin" and run.get("created_by") != current_user["id"]:
        raise HTTPException(403, "Access denied")
    if run.get("status") not in ("processed", "exported"):
        raise HTTPException(409, "Run has not been processed yet — process it first")
    return run


def _check_integrity(run_id: str, workflow_type: str) -> None:
    """PRD §7.3/§16.1 — block export if blocking validation failed and no override."""
    if workflow_type == "jetro_reconciliation":
        invoices = result_store.load(run_id, "jetro_invoices", [])
        bad = [i for i in invoices if i.get("integrity_status") != "ok"]
        if bad and not run_store.has_override(run_id, "jetro_integrity_mismatch"):
            invoice_nums = ", ".join(str(i.get("invoice_number")) for i in bad)
            raise HTTPException(
                409,
                detail={
                    "error": "integrity_mismatch",
                    "message": (
                        f"Grand total mismatch on invoice(s) {invoice_nums}. "
                        "Resolve the discrepancy or record an authorized override "
                        "(POST /runs/{run_id}/override with check='jetro_integrity_mismatch') before exporting."
                    ),
                    "invoices": [i.get("invoice_number") for i in bad],
                },
            )


def _send(file_path: str, filename: str) -> FastAPIFileResponse:
    p = Path(file_path)
    if not p.exists():
        raise HTTPException(500, "Export file not found on disk — try re-exporting")
    return FastAPIFileResponse(path=str(p), filename=filename, media_type=XLSX_MEDIA)


@router.post("/web-orders/{run_id}")
async def export_web_orders(
    run_id: str,
    current_user: dict = Depends(require_roles(*_EXPORT_ROLES)),
):
    run = _resolve(run_id, "web_orders_check", current_user)
    _check_integrity(run_id, "web_orders_check")
    file_path = await wo_service.export_workbook(run_id)
    run_store.update_run(run_id, {"status": "exported"})
    audit_store.log(run_id, "exported", "Web orders workbook exported", current_user["id"])
    date_str = (run.get("run_date") or "").replace("-", "")
    return _send(file_path, f"Web_Orders_Check_{date_str}.xlsx")


@router.post("/jetro/{run_id}")
async def export_jetro(
    run_id: str,
    current_user: dict = Depends(require_roles(*_EXPORT_ROLES)),
):
    run = _resolve(run_id, "jetro_reconciliation", current_user)
    _check_integrity(run_id, "jetro_reconciliation")
    file_path = await jetro_service.export_workbook(run_id)
    run_store.update_run(run_id, {"status": "exported"})
    audit_store.log(run_id, "exported", "Jetro workbook exported", current_user["id"])
    date_str = (run.get("run_date") or "").replace("-", "")
    return _send(file_path, f"Jetro_Restaurant_Depot_Reconciliation_{date_str}.xlsx")


@router.post("/vendor-bills/{run_id}")
async def export_vendor_bills(
    run_id: str,
    current_user: dict = Depends(require_roles(*_EXPORT_ROLES)),
):
    run = _resolve(run_id, "vendor_bill_po_bank", current_user)
    _check_integrity(run_id, "vendor_bill_po_bank")
    file_path = await vb_service.export_workbook(run_id)
    run_store.update_run(run_id, {"status": "exported"})
    audit_store.log(run_id, "exported", "Vendor bills workbook exported", current_user["id"])
    date_str = (run.get("run_date") or "").replace("-", "")
    return _send(file_path, f"QB_Vendor_Bill_Import_{date_str}_Accumulating.xlsx")


@router.post("/vendor-bills/{run_id}/draft")
async def export_vendor_bills_draft(
    run_id: str,
    current_user: dict = Depends(require_roles(*_EXPORT_ROLES)),
):
    """Draft export — generates a Bill Import workbook from currently confirmed
    bills WITHOUT running the unbilled-PO sweep and WITHOUT marking the run as
    exported. Used while bills are still trickling in from vendors so they're
    not prematurely flagged as missing. Finalize remains a separate action."""
    run = _resolve(run_id, "vendor_bill_po_bank", current_user)
    _check_integrity(run_id, "vendor_bill_po_bank")
    file_path = await vb_service.export_workbook(run_id, finalize=False)
    audit_store.log(run_id, "exported_draft",
                    "Vendor bills draft workbook exported (no PO sweep)",
                    current_user["id"])
    date_str = (run.get("run_date") or "").replace("-", "")
    return _send(file_path, f"QB_Vendor_Bill_Import_{date_str}_Draft.xlsx")


@router.post("/combined-price/{run_id}")
async def export_combined_price(
    run_id: str,
    current_user: dict = Depends(require_roles(*_EXPORT_ROLES)),
):
    run = _resolve(run_id, "combined_price_changes", current_user)
    _check_integrity(run_id, "combined_price_changes")
    file_path = await cp_service.export_workbook(run_id)
    run_store.update_run(run_id, {"status": "exported"})
    audit_store.log(run_id, "exported", "Combined price workbook exported", current_user["id"])
    date_str = (run.get("run_date") or "").replace("-", "")
    return _send(file_path, f"Price_Changes_Both_Sources_{date_str}.xlsx")


# ---------------------------------------------------------------------------
# Sample template download — pre-built XLSX with correct headers / columns
# ---------------------------------------------------------------------------

_SAMPLE_TEMPLATES: dict[str, tuple[str, list]] = {
    # QB PO export (cols A-M) ------------------------------------------------
    # Columns derived from vendor_bills/parser.py classify_po_row + product_lines
    "qb-po-export": (
        "QB_PO_Export_Template.xlsx",
        [
            "Terms",
            "Ref #",
            "TxnDate",
            "Vendor",
            "Memo",
            "TxnLine Amount",
            "TxnLine Cost",
            "TxnLine Description",
            "TxnLine Item",
            "TxnLine Quantity",
            "Class",
            "Case Avg Weight",
            "Unit Avg Weight",
        ],
    ),
    # Web Orders master spreadsheet (All Orders sheet, cols A-S) -------------
    # Columns derived from web_orders/parser.py parse_all_orders
    "web-orders-master": (
        "Web_Orders_Master_Template.xlsx",
        [
            "Product Name",          # A
            "Code",                  # B
            "Col C",                 # C (passthrough)
            "Col D",                 # D (passthrough)
            "Col E",                 # E (passthrough)
            "Col F",                 # F (passthrough)
            "Qty",                   # G
            "Weight",                # H
            "Individual Weight Status",  # I
            "Price",                 # J
            "Col K",                 # K (passthrough)
            "Customer Name",         # L
            "Col M",                 # M (passthrough)
            "Col N",                 # N (passthrough)
            "Transaction Date",      # O
            "Col P",                 # P (passthrough)
            "Route",                 # Q
            "Col R",                 # R (passthrough)
            "Remark",                # S
        ],
    ),
    # Jetro source sheet (cols A-AI key ones) ---------------------------------
    # Columns derived from jetro/parser.py parse_jetro_source
    "jetro-source": (
        "Jetro_Source_Template.xlsx",
        [
            "Col A",           # A  (index 0 — unused by parser)
            "Col B",           # B  (index 1 — unused)
            "Qty",             # C  (index 2)
            "Product Name",    # D  (index 3)
            "Item Code",       # E  (index 4)  — append "U" suffix for unit items
            "Col F",           # F  (indices 5-13 passthrough)
            "Col G", "Col H", "Col I", "Col J", "Col K", "Col L", "Col M",
            "Col N",           # N  (index 13)
            "Customer",        # O  (index 14)
            "Col P", "Col Q", "Col R", "Col S", "Col T", "Col U", "Col V", "Col W",  # P-W
            "Current Cost Price",    # X  (index 23)
            "Current Selling Price", # Y  (index 24)
            "Col Z",           # Z  (index 25)
            "Col AA", "Col AB", "Col AC", "Col AD", "Col AE",  # AA-AE
            "Case Avg Weight", # AF (index 31)
            "Col AG", "Col AH",  # AG-AH
            "Unit",            # AI (index 34)
        ],
    ),
}


@router.get("/sample-template/{template}")
async def download_sample_template(
    template: str,
    current_user: dict = Depends(require_roles(*_EXPORT_ROLES)),
):
    """Stream a pre-built XLSX sample with the correct column headers for the
    given template name. Use as a formatting guide before uploading files."""
    entry = _SAMPLE_TEMPLATES.get(template)
    if not entry:
        raise HTTPException(404, f"Unknown template '{template}'. "
                            f"Valid options: {', '.join(_SAMPLE_TEMPLATES)}")
    filename, headers = entry

    buf = io.BytesIO()
    from openpyxl import Workbook as _WB
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = _WB()
    ws = wb.active
    ws.title = "Template"
    ws.append(headers)
    # Style header row
    header_fill = PatternFill("solid", fgColor="1F4E79")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        ws.column_dimensions[cell.column_letter].width = max(len(str(cell.value or "")) + 4, 14)
    ws.row_dimensions[1].height = 20
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type=XLSX_MEDIA,
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


_WORKFLOW_EXPORTERS = {
    "web_orders_check": (wo_service.export_workbook, "Web_Orders_Check"),
    "jetro_reconciliation": (jetro_service.export_workbook, "Jetro_Restaurant_Depot_Reconciliation"),
    "vendor_bill_po_bank": (vb_service.export_workbook, "QB_Vendor_Bill_Import"),
    "combined_price_changes": (cp_service.export_workbook, "Price_Changes_Both_Sources"),
}


@router.post("/day-archive/{date}")
async def export_day_archive(
    date: str,
    current_user: dict = Depends(require_roles(*_EXPORT_ROLES)),
):
    """Zip archive of the most recent finalized export for each workflow on a
    given calendar day (YYYY-MM-DD). Only runs in `exported` state are
    eligible — i.e. the user has actually exported the final workbook at least
    once. For each workflow_type only the most recently updated qualifying run
    is included. Filename is `<date>.zip`."""
    all_runs = run_store.list_runs()
    # Filter to the requested day, eligible status, and (non-admin) ownership.
    eligible = []
    for r in all_runs:
        if not (r.get("run_date") or "").startswith(date):
            continue
        if r.get("status") != "exported":
            continue
        if current_user.get("role") != "admin" and r.get("created_by") != current_user["id"]:
            continue
        if r.get("workflow_type") not in _WORKFLOW_EXPORTERS:
            continue
        eligible.append(r)

    if not eligible:
        raise HTTPException(404, "No finalized runs found for that date")

    # Most recent per workflow_type by updated_at.
    latest_by_type: dict[str, dict] = {}
    for r in eligible:
        wt = r["workflow_type"]
        prev = latest_by_type.get(wt)
        if prev is None or (r.get("updated_at") or "") > (prev.get("updated_at") or ""):
            latest_by_type[wt] = r

    buf = io.BytesIO()
    date_compact = date.replace("-", "")
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for wt, run in latest_by_type.items():
            exporter, prefix = _WORKFLOW_EXPORTERS[wt]
            try:
                _check_integrity(run["id"], wt)
            except HTTPException:
                # Skip runs that still have unresolved blocking issues.
                continue
            file_path = await exporter(run["id"])
            p = Path(file_path)
            if not p.exists():
                continue
            zf.write(p, arcname=f"{prefix}_{date_compact}.xlsx")

    if not buf.getvalue() or len(buf.getvalue()) < 30:
        raise HTTPException(404, "No exportable runs found for that date")

    buf.seek(0)
    audit_store.log("system", "day_archive_exported",
                    f"Day archive exported for {date} ({len(latest_by_type)} workflow(s))",
                    current_user["id"])
    return StreamingResponse(
        buf,
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="Morning Report {date}.zip"'},
    )
